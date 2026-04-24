from __future__ import annotations
from copy import copy
from datetime import date, datetime, time
from pathlib import Path
import io, shutil

from openpyxl import load_workbook
from openpyxl.styles import Font

from .data_models import TravelFormData


class TravelTemplateError(RuntimeError):
    pass


class TravelTemplateService:
    """
    Compila il template Trasferta_template.xlsx con openpyxl,
    preservando tutto lo stile e la struttura originale.
    Stessa logica del WeeklyProgramTemplateService.
    """

    TRANSFER_ROWS = [18, 19, 20, 21, 22, 23, 24, 25]
    STAY_ROWS     = [36, 37, 38]
    BLACK         = 'FF000000'

    def __init__(self, template_path: str | Path):
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(
                f'Template non trovato: {self.template_path}\n'
                'Salva il file "NEW - Mod. Trasferta.xls" come '
                '"Trasferta_template.xlsx" nella cartella templates_excel/.'
            )

    # ── Generazione in memoria ────────────────────────────────────────────────
    def generate_bytes(self, data: TravelFormData) -> bytes:
        wb = load_workbook(self.template_path)
        ws = wb.active

        self._write_travel_type(ws, data.travel_type)
        self._set(ws, 'F5', self._d(data.start_date))
        self._set(ws, 'I5', self._d(data.end_date))
        self._set(ws, 'C8', data.full_name)
        self._set(ws, 'I8', data.employee_id)
        self._set(ws, 'C9', data.business_line)
        self._set(ws, 'I9', data.cost_center)
        self._set(ws, 'C10', data.hiring_location)
        self._set(ws, 'C11', data.travel_reason)

        # Svuota tutte le righe trasferimento
        for row_n in self.TRANSFER_ROWS:
            for col in ('A', 'C', 'E', 'F', 'G', 'I'):
                self._set(ws, f'{col}{row_n}', '')

        # Scrive i trasferimenti
        for row_n, tr in zip(self.TRANSFER_ROWS, data.transfers):
            self._set(ws, f'A{row_n}', tr.from_location)
            self._set(ws, f'C{row_n}', tr.to_location)
            self._set(ws, f'E{row_n}', self._d(tr.travel_date))
            self._set(ws, f'F{row_n}', self._t(tr.departure_time))
            self._set(ws, f'G{row_n}', tr.transport)
            self._set(ws, f'I{row_n}', tr.notes)

        # Noleggio
        r = data.rental
        self._set(ws, 'B30', r.pickup_location)
        self._set(ws, 'F30', self._d(r.pickup_date))
        self._set(ws, 'I30', self._t(r.pickup_time))
        self._set(ws, 'B31', r.dropoff_location)
        self._set(ws, 'F31', self._d(r.dropoff_date))
        self._set(ws, 'I31', self._t(r.dropoff_time))
        self._set(ws, 'B32', r.vehicle_category)

        # Svuota pernottamenti
        for row_n in self.STAY_ROWS:
            for col in ('B', 'F', 'I'):
                self._set(ws, f'{col}{row_n}', '')

        for row_n, st in zip(self.STAY_ROWS, data.stays):
            self._set(ws, f'B{row_n}', st.location)
            self._set(ws, f'F{row_n}', self._d(st.check_in))
            self._set(ws, f'I{row_n}', self._d(st.check_out))

        # Firma
        self._set(ws, 'C43', data.employee_signature_name or data.full_name)
        self._set(ws, 'I43', self._d(data.employee_signature_date))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ── Helper ────────────────────────────────────────────────────────────────
    def _set(self, ws, cell_ref: str, value) -> None:
        """
        Scrive il valore nella cella, gestendo le merged cells.
        openpyxl non permette di scrivere su MergedCell (celle non-master):
        in quel caso cerca la cella master del range e scrive lì.
        """
        from openpyxl.cell import MergedCell
        c = ws[cell_ref]
        if isinstance(c, MergedCell):
            # Cerca il range di merge che contiene questa cella e usa la master
            for rng in ws.merged_cells.ranges:
                if cell_ref in rng:
                    master = ws.cell(rng.min_row, rng.min_col)
                    master.value = value
                    try:
                        f = copy(master.font)
                        f.color = self.BLACK
                        master.font = f
                    except Exception:
                        pass
                    return
            return  # merged senza range trovato — skip silenzioso
        c.value = value
        try:
            f = copy(c.font)
            f.color = self.BLACK
            c.font = f
        except Exception:
            pass

    def _write_travel_type(self, ws, travel_type: str) -> None:
        mapping = {
            'forfettaria': 'A4',
            'piè di lista': 'A5',
            'pie di lista': 'A5',
            'mista': 'A6',
        }
        for cell in ('A4', 'A5', 'A6'):
            self._set(ws, cell, '☐')
        target = mapping.get((travel_type or '').lower(), 'A6')
        self._set(ws, target, '☑')

    @staticmethod
    def _d(v: date | None):
        if not v:
            return ''
        return datetime(v.year, v.month, v.day)

    @staticmethod
    def _t(v: time | None) -> str:
        return v.strftime('%H:%M') if v else ''
