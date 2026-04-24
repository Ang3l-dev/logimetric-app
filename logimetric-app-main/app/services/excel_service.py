from __future__ import annotations
from datetime import date, datetime, time
from pathlib import Path
import io, shutil

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

from .data_models import TravelFormData

try:
    import pythoncom       # type: ignore
    import win32com.client as win32  # type: ignore
    _WIN32 = True
except Exception:
    _WIN32 = False


class ExcelGenerationError(RuntimeError):
    pass


# ─────────────────────────────────────────────────────────────────────────────
# Servizio via COM (Windows + Excel installato)
# ─────────────────────────────────────────────────────────────────────────────
class ExcelTemplateService:
    """Usa COM automation su Windows per compilare il template .xls originale."""
    TRANSFER_ROWS = [18, 19, 20, 21, 22, 23, 24, 25]
    STAY_ROWS     = [36, 37, 38]

    def __init__(self, template_path: str | Path):
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f'Template non trovato: {self.template_path}')

    @staticmethod
    def _xl_date(v: date | None):
        return datetime(v.year, v.month, v.day) if v else ''

    @staticmethod
    def _xl_time(v: time | None) -> str:
        return v.strftime('%H:%M') if v else ''

    def generate(self, data: TravelFormData, target_dir: str | Path) -> dict[str, str]:
        if not _WIN32:
            raise ExcelGenerationError(
                'win32com non disponibile. Usa la generazione openpyxl (Linux/cloud).')
        target_dir = Path(target_dir)
        target_dir.mkdir(parents=True, exist_ok=True)
        out_xls = target_dir / f'{data.safe_filename_base}.xls'
        out_pdf = target_dir / f'{data.safe_filename_base}.pdf'
        shutil.copy2(self.template_path, out_xls)

        pythoncom.CoInitialize()
        excel = wb = None
        try:
            excel = win32.DispatchEx('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(out_xls.resolve()))
            ws = wb.Worksheets(1)

            self._write_checkbox(ws, data.travel_type)
            ws.Range('F5').Value = self._xl_date(data.start_date)
            ws.Range('I5').Value = self._xl_date(data.end_date)
            ws.Range('C8').Value = data.full_name
            ws.Range('I8').Value = data.employee_id
            ws.Range('C9').Value = data.business_line
            ws.Range('I9').Value = data.cost_center
            ws.Range('C10').Value = data.hiring_location
            ws.Range('C11').Value = data.travel_reason

            for row_n in self.TRANSFER_ROWS:
                for col in ('A','C','E','F','G','I'):
                    ws.Range(f'{col}{row_n}').Value = ''
            for row_n, tr in zip(self.TRANSFER_ROWS, data.transfers):
                ws.Range(f'A{row_n}').Value = tr.from_location
                ws.Range(f'C{row_n}').Value = tr.to_location
                ws.Range(f'E{row_n}').Value = self._xl_date(tr.travel_date)
                ws.Range(f'F{row_n}').Value = self._xl_time(tr.departure_time)
                ws.Range(f'G{row_n}').Value = tr.transport
                ws.Range(f'I{row_n}').Value = tr.notes

            ws.Range('B30').Value = data.rental.pickup_location
            ws.Range('F30').Value = self._xl_date(data.rental.pickup_date)
            ws.Range('I30').Value = self._xl_time(data.rental.pickup_time)
            ws.Range('B31').Value = data.rental.dropoff_location
            ws.Range('F31').Value = self._xl_date(data.rental.dropoff_date)
            ws.Range('I31').Value = self._xl_time(data.rental.dropoff_time)
            ws.Range('B32').Value = data.rental.vehicle_category

            for row_n in self.STAY_ROWS:
                for col in ('B','F','I'):
                    ws.Range(f'{col}{row_n}').Value = ''
            for row_n, st in zip(self.STAY_ROWS, data.stays):
                ws.Range(f'B{row_n}').Value = st.location
                ws.Range(f'F{row_n}').Value = self._xl_date(st.check_in)
                ws.Range(f'I{row_n}').Value = self._xl_date(st.check_out)

            ws.Range('C43').Value = data.employee_signature_name or data.full_name
            ws.Range('I43').Value = self._xl_date(data.employee_signature_date)

            wb.Save()
            wb.ExportAsFixedFormat(0, str(out_pdf.resolve()))
        except Exception as exc:
            raise ExcelGenerationError(str(exc)) from exc
        finally:
            if wb:    wb.Close(SaveChanges=True)
            if excel: excel.Quit()
            pythoncom.CoUninitialize()

        return {'xls': str(out_xls), 'pdf': str(out_pdf)}

    def _write_checkbox(self, ws, travel_type: str) -> None:
        mapping = {'forfettaria': 'A4', 'piè di lista': 'A5', 'pie di lista': 'A5', 'mista': 'A6'}
        for c in ('A4','A5','A6'):
            ws.Range(c).Value = '☐'
        ws.Range(mapping.get((travel_type or '').lower(), 'A6')).Value = '☑'


# ─────────────────────────────────────────────────────────────────────────────
# Generatore openpyxl (Linux / Render)
# ─────────────────────────────────────────────────────────────────────────────
class TravelFormXlsxGenerator:
    """Genera un .xlsx ben formattato senza dipendenze Windows."""

    # Palette colori
    C_HEADER    = '0D1830'
    C_ACCENT    = '1FA3FF'
    C_SECTION   = '123B6D'
    C_LABEL_BG  = 'EDF3FB'
    C_WHITE     = 'FFFFFF'
    C_BORDER    = 'D0DCF0'

    def generate(self, data: TravelFormData) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Modulo Trasferta'
        ws.sheet_view.showGridLines = False

        # Larghezze colonne
        col_widths = [4, 22, 18, 14, 14, 14, 16, 16, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = 1

        # ── Intestazione ──────────────────────────────────────────
        row = self._header_block(ws, row, data)

        # ── Dati principali ────────────────────────────────────────
        row = self._section_title(ws, row, 'DATI PRINCIPALI')
        fields_main = [
            ('Tipo trasferta',   data.travel_type.capitalize()),
            ('Dal',              data.start_date.strftime('%d/%m/%Y') if data.start_date else ''),
            ('Al',               data.end_date.strftime('%d/%m/%Y') if data.end_date else ''),
            ('Cognome e Nome',   data.full_name),
            ('Matricola',        data.employee_id),
            ('Linea di business',data.business_line),
            ('Centro di costo',  data.cost_center),
            ('Sede di assunzione',data.hiring_location),
            ('Motivo del viaggio',data.travel_reason),
        ]
        for label, value in fields_main:
            row = self._field_row(ws, row, label, value)

        # ── Trasferimenti ──────────────────────────────────────────
        row += 1
        row = self._section_title(ws, row, 'TRASFERIMENTI')
        headers_tr = ['Da', 'A', 'Data', 'Orario', 'Mezzo', 'Note']
        row = self._table_header(ws, row, headers_tr, col_span=2)
        for tr in data.transfers:
            vals = [
                tr.from_location,
                tr.to_location,
                tr.travel_date.strftime('%d/%m/%Y') if tr.travel_date else '',
                tr.departure_time.strftime('%H:%M') if tr.departure_time else '',
                tr.transport,
                tr.notes,
            ]
            row = self._table_row(ws, row, vals)

        # ── Noleggio auto ──────────────────────────────────────────
        row += 1
        row = self._section_title(ws, row, 'NOLEGGIO AUTO')
        r = data.rental
        rental_fields = [
            ('Ritiro — Luogo',  r.pickup_location),
            ('Ritiro — Data',   r.pickup_date.strftime('%d/%m/%Y') if r.pickup_date else ''),
            ('Ritiro — Orario', r.pickup_time.strftime('%H:%M') if r.pickup_time else ''),
            ('Riconsegna — Luogo',  r.dropoff_location),
            ('Riconsegna — Data',   r.dropoff_date.strftime('%d/%m/%Y') if r.dropoff_date else ''),
            ('Riconsegna — Orario', r.dropoff_time.strftime('%H:%M') if r.dropoff_time else ''),
            ('Categoria veicolo', r.vehicle_category),
        ]
        for label, value in rental_fields:
            row = self._field_row(ws, row, label, value)

        # ── Pernottamenti ──────────────────────────────────────────
        row += 1
        row = self._section_title(ws, row, 'PERNOTTAMENTI')
        headers_st = ['Città/Hotel', 'Check-in', 'Check-out']
        row = self._table_header(ws, row, headers_st, col_span=3)
        for st in data.stays:
            vals = [
                st.location,
                st.check_in.strftime('%d/%m/%Y') if st.check_in else '',
                st.check_out.strftime('%d/%m/%Y') if st.check_out else '',
            ]
            row = self._table_row(ws, row, vals)

        # ── Firma ─────────────────────────────────────────────────
        row += 1
        row = self._section_title(ws, row, 'FIRMA DIPENDENTE')
        row = self._field_row(ws, row, 'Nome firma', data.employee_signature_name or data.full_name)
        row = self._field_row(ws, row, 'Data firma',
                              data.employee_signature_date.strftime('%d/%m/%Y')
                              if data.employee_signature_date else '')

        # ── Scrivi in memoria ──────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ── Helper di stile ───────────────────────────────────────────
    def _hfont(self, bold=False, color='000000', size=11):
        return Font(name='Calibri', bold=bold, color=color, size=size)

    def _fill(self, hex_color: str):
        return PatternFill('solid', fgColor=hex_color)

    def _thin_border(self):
        s = Side(style='thin', color=self.C_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def _header_block(self, ws, row: int, data: TravelFormData) -> int:
        ws.merge_cells(f'A{row}:I{row}')
        c = ws.cell(row, 1, 'MODULO TRASFERTA — LogiMetric')
        c.font = Font(name='Calibri', bold=True, color=self.C_WHITE, size=16)
        c.fill = self._fill(self.C_HEADER)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 36
        row += 1

        ws.merge_cells(f'A{row}:I{row}')
        c = ws.cell(row, 1, f'Generato il {datetime.now().strftime("%d/%m/%Y %H:%M")}')
        c.font = Font(name='Calibri', color='AABBCC', size=10, italic=True)
        c.fill = self._fill(self.C_SECTION)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 20
        return row + 2

    def _section_title(self, ws, row: int, title: str) -> int:
        ws.merge_cells(f'A{row}:I{row}')
        c = ws.cell(row, 1, title)
        c.font = Font(name='Calibri', bold=True, color=self.C_WHITE, size=11)
        c.fill = self._fill(self.C_ACCENT)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 22
        return row + 1

    def _field_row(self, ws, row: int, label: str, value: str) -> int:
        lc = ws.cell(row, 2, label)
        lc.font = self._hfont(bold=True, color='1D2A44')
        lc.fill = self._fill(self.C_LABEL_BG)
        lc.border = self._thin_border()
        lc.alignment = Alignment(vertical='center', indent=1)

        ws.merge_cells(f'C{row}:I{row}')
        vc = ws.cell(row, 3, value)
        vc.font = self._hfont(color='1D2A44')
        vc.border = self._thin_border()
        vc.alignment = Alignment(vertical='center', wrap_text=True, indent=1)
        ws.row_dimensions[row].height = 18
        return row + 1

    def _table_header(self, ws, row: int, headers: list[str], col_span: int = 1) -> int:
        col = 2
        for h in headers:
            end_col = col + col_span - 1
            if col_span > 1:
                ws.merge_cells(start_row=row, start_column=col,
                               end_row=row, end_column=end_col)
            c = ws.cell(row, col, h)
            c.font = Font(name='Calibri', bold=True, color=self.C_WHITE, size=10)
            c.fill = self._fill(self.C_SECTION)
            c.border = self._thin_border()
            c.alignment = Alignment(horizontal='center', vertical='center')
            col = end_col + 1
        ws.row_dimensions[row].height = 18
        return row + 1

    def _table_row(self, ws, row: int, values: list[str]) -> int:
        fill = self._fill('F4F8FF') if row % 2 == 0 else self._fill(self.C_WHITE)
        for i, val in enumerate(values):
            c = ws.cell(row, i + 2, val)
            c.font = self._hfont(color='1D2A44', size=10)
            c.fill = fill
            c.border = self._thin_border()
            c.alignment = Alignment(vertical='center', indent=1)
        ws.row_dimensions[row].height = 17
        return row + 1
