from __future__ import annotations
import io
from datetime import date, timedelta, datetime

from flask import (abort, current_app, flash, jsonify, redirect,
                   render_template, request, send_file, url_for)
from flask_login import current_user, login_required

from . import aste_bp
from .. import db
from ..models import (House, HouseParticipant, HouseExpense,
                      HouseDeadline, HouseAudit, User,
                      HOUSE_STATI, HOUSE_STATI_COLORS, EXPENSE_CATS)


def _require_view():
    if not current_user.is_authenticated or not current_user.can_view('gestione_aste'):
        abort(403)

def _require_write():
    if not current_user.is_authenticated or not current_user.can_write('gestione_aste'):
        abort(403)


# ── Dashboard / griglia ───────────────────────────────────────────────────────
@aste_bp.get('')
@login_required
def index():
    _require_view()
    houses = House.query.order_by(House.data_asta.desc().nullslast(),
                                  House.created_at.desc()).all()
    users  = User.query.filter_by(is_approved=True).order_by(User.name).all()
    kpi = _compute_kpi(houses)
    return render_template('aste/index.html',
                           houses=houses, users=users, kpi=kpi,
                           stati=HOUSE_STATI, stati_colors=HOUSE_STATI_COLORS,
                           expense_cats=EXPENSE_CATS)


# ── Scheda personale ──────────────────────────────────────────────────────────
@aste_bp.get('/la-mia-scheda')
@login_required
def mia_scheda():
    _require_view()
    """Vista personale: per ogni casa a cui partecipo, il mio saldo."""
    participations = HouseParticipant.query.filter_by(user_id=current_user.id).all()
    house_ids = [p.house_id for p in participations]
    houses = House.query.filter(House.id.in_(house_ids)).all() if house_ids else []

    rows = []
    tot_pagato        = 0
    tot_quota         = 0
    tot_saldo_display = 0

    for h in houses:
        sal  = h.saldo_utente(current_user.id)
        part = h.participants.filter_by(user_id=current_user.id).first()
            # Dettaglio per ogni spesa
        exp_rows = []
        for exp in h.expenses.order_by(HouseExpense.data.desc()).all():
            from ..models import HouseExpenseSettlement
            is_payer  = (exp.paid_by_user_id == current_user.id)
            quota_exp = (exp.importo or 0) / h.n_partecipanti

            if is_payer:
                debtors_status = []
                for p2 in h.participants.all():
                    if p2.user_id == current_user.id:
                        continue
                    stl = HouseExpenseSettlement.query.filter_by(
                        expense_id=exp.id, user_id=p2.user_id).first()
                    debtors_status.append({'user': p2.user, 'versato': stl.versato if stl else False, 'quota': quota_exp})
                creditore_stl = HouseExpenseSettlement.query.filter_by(
                    expense_id=exp.id, user_id=current_user.id).first()
                exp_rows.append({
                    'expense': exp, 'is_payer': True, 'quota': quota_exp,
                    'debtors_status': debtors_status,
                    'ricevuto': creditore_stl.ricevuto if creditore_stl else False,
                    'tutti_versato': all(d['versato'] for d in debtors_status) if debtors_status else False,
                })
            else:
                stl = HouseExpenseSettlement.query.filter_by(
                    expense_id=exp.id, user_id=current_user.id).first()
                exp_rows.append({
                    'expense': exp, 'is_payer': False, 'quota': quota_exp,
                    'versato': stl.versato if stl else False,
                    'versato_il': stl.versato_il if stl else None,
                    'paid_by': exp.paid_by,
                })

        rows.append({
            'house':            h,
            'quota':            sal['quota'],
            'pagato':           sal['pagato'],
            'saldo':            sal['saldo'],
            'saldo_display':    sal['saldo_display'],
            'deve_ricevere':    sal['deve_ricevere'],
            'deve_dare':        sal['deve_dare'],
            'quota_versata':    sal['quota_versata'],
            'credito_ricevuto': sal['credito_ricevuto'],
            'liquidato':        sal['liquidato'],
            'versato_il':       part.versato_il  if part else None,
            'ricevuto_il':      part.ricevuto_il if part else None,
            'exp_rows':         exp_rows,
        })
        tot_pagato        += sal['pagato']
        tot_quota         += sal['quota']
        tot_saldo_display += sal['saldo_display']

    return render_template('aste/mia_scheda.html',
                           rows=rows,
                           tot_pagato=tot_pagato,
                           tot_quota=tot_quota,
                           tot_saldo=tot_saldo_display,
                           stati_colors=HOUSE_STATI_COLORS,
                           stati=HOUSE_STATI)


# ── Conferma versamento quota ──────────────────────────────────────────────────
@aste_bp.post('/<int:house_id>/conferma-versamento')
@login_required
def conferma_versamento(house_id: int):
    """Debitore: conferma di aver versato la sua quota."""
    _require_write()
    part = HouseParticipant.query.filter_by(
        house_id=house_id, user_id=current_user.id).first_or_404()
    part.quota_versata = not part.quota_versata
    part.versato_il    = datetime.utcnow() if part.quota_versata else None
    db.session.commit()
    return jsonify({'ok': True, 'versata': part.quota_versata})


@aste_bp.post('/<int:house_id>/conferma-ricezione')
@login_required
def conferma_ricezione(house_id: int):
    """Creditore: conferma di aver ricevuto le quote dai debitori."""
    _require_write()
    part  = HouseParticipant.query.filter_by(
        house_id=house_id, user_id=current_user.id).first_or_404()
    house = House.query.get_or_404(house_id)
    saldo = house.saldo_utente(current_user.id)

    if not saldo['deve_ricevere']:
        return jsonify({'ok': False, 'error': 'Non hai un credito su questo immobile.'}), 400

    part.credito_ricevuto = not part.credito_ricevuto
    part.ricevuto_il      = datetime.utcnow() if part.credito_ricevuto else None
    db.session.commit()
    return jsonify({'ok': True, 'ricevuto': part.credito_ricevuto})


# ── Nuova casa ────────────────────────────────────────────────────────────────
@aste_bp.get('/new')
@login_required
def new_house():
    _require_write()
    users = User.query.filter_by(is_approved=True).order_by(User.name).all()
    return render_template('aste/form.html', house=None, users=users,
                           stati=HOUSE_STATI, expense_cats=EXPENSE_CATS)


@aste_bp.post('/new')
@login_required
def create_house():
    _require_write()
    house = _house_from_form(None)
    db.session.add(house)
    db.session.flush()
    _save_participants(house)
    _save_expenses(house)
    _save_deadlines(house)
    _audit(house, 'creazione', '', f'Casa creata: {house.via}, {house.citta}')
    db.session.commit()
    flash(f'Immobile "{house.via}, {house.citta}" aggiunto.', 'success')
    return redirect(url_for('aste.index'))


# ── Dettaglio ─────────────────────────────────────────────────────────────────
@aste_bp.get('/<int:house_id>')
@login_required
def detail(house_id: int):
    _require_view()
    from ..models import HouseExpenseSettlement
    house     = House.query.get_or_404(house_id)
    users     = User.query.filter_by(is_approved=True).order_by(User.name).all()
    part_ids  = {p.user_id for p in house.participants.all()}
    parts     = house.participants.all()
    expenses  = house.expenses.order_by(HouseExpense.data.desc()).all()
    deadlines = house.deadlines.order_by(HouseDeadline.data_scadenza).all()
    audit     = house.audit_log.order_by(HouseAudit.modified_at.desc()).limit(30).all()
    n         = house.n_partecipanti

    saldi = []
    for p in parts:
        sal = house.saldo_utente(p.user_id)
        saldi.append({'participant': p, 'user': p.user, **sal})

    # Per ogni spesa: chi deve dare, chi deve ricevere, stato saldo
    exp_details = []
    for exp in expenses:
        quota   = (exp.importo or 0) / n
        debtors = []
        for p in parts:
            if p.user_id == exp.paid_by_user_id:
                continue  # chi ha pagato non è debitore su questa spesa
            stl = HouseExpenseSettlement.query.filter_by(
                expense_id=exp.id, user_id=p.user_id).first()
            debtors.append({
                'user':     p.user,
                'quota':    quota,
                'versato':  stl.versato   if stl else False,
                'versato_il': stl.versato_il if stl else None,
            })
        # Creditore (chi ha pagato)
        creditore_stl = HouseExpenseSettlement.query.filter_by(
            expense_id=exp.id, user_id=exp.paid_by_user_id).first() if exp.paid_by_user_id else None
        exp_details.append({
            'expense':    exp,
            'quota':      quota,
            'debtors':    debtors,
            'ricevuto':   creditore_stl.ricevuto    if creditore_stl else False,
            'ricevuto_il': creditore_stl.ricevuto_il if creditore_stl else None,
        })

    return render_template('aste/detail.html',
                           house=house, users=users, part_ids=part_ids,
                           parts=parts, saldi=saldi,
                           exp_details=exp_details,
                           deadlines=deadlines, audit=audit,
                           stati=HOUSE_STATI, stati_colors=HOUSE_STATI_COLORS,
                           expense_cats=EXPENSE_CATS)


# ── Modifica ──────────────────────────────────────────────────────────────────
@aste_bp.get('/<int:house_id>/edit')
@login_required
def edit_house(house_id: int):
    _require_write()
    house = House.query.get_or_404(house_id)
    users = User.query.filter_by(is_approved=True).order_by(User.name).all()
    return render_template('aste/form.html', house=house, users=users,
                           stati=HOUSE_STATI, expense_cats=EXPENSE_CATS)


@aste_bp.post('/<int:house_id>/edit')
@login_required
def update_house(house_id: int):
    _require_write()
    house = House.query.get_or_404(house_id)
    _update_from_form(house)
    _save_participants(house)
    db.session.commit()
    flash('Immobile aggiornato.', 'success')
    return redirect(url_for('aste.detail', house_id=house.id))


# ── Elimina ───────────────────────────────────────────────────────────────────
@aste_bp.post('/<int:house_id>/delete')
@login_required
def delete_house(house_id: int):
    _require_write()
    house = House.query.get_or_404(house_id)
    label = f'{house.via}, {house.citta}'
    db.session.delete(house)
    db.session.commit()
    flash(f'Immobile "{label}" eliminato.', 'info')
    return redirect(url_for('aste.index'))


@aste_bp.post('/<int:house_id>/registra-vendita')
@login_required
def registra_vendita(house_id: int):
    """Registra il prezzo di vendita effettivo e calcola i guadagni."""
    _require_write()
    house = House.query.get_or_404(house_id)

    prezzo_vendita = _to_float(request.form.get('prezzo_vendita'))
    if prezzo_vendita <= 0:
        flash('Inserisci un prezzo di vendita valido.', 'error')
        return redirect(url_for('aste.detail', house_id=house_id))

    old_stato = house.stato
    house.ricavo_consuntivo = prezzo_vendita
    house.stato             = 'venduta'

    _audit(house, 'vendita',
           f'stato={old_stato}, ricavo={house.ricavo_consuntivo}',
           f'venduta a {prezzo_vendita} € — guadagno netto: {house.guadagno_reale:.2f} €')
    db.session.commit()

    flash(
        f'Vendita registrata: {prezzo_vendita:,.2f} € · '
        f'Guadagno netto: {house.guadagno_reale:,.2f} € · '
        f'Quota per partecipante: {house.guadagno_reale / house.n_partecipanti:,.2f} €'
        .replace(',', 'X').replace('.', ',').replace('X', '.'),
        'success'
    )
    return redirect(url_for('aste.detail', house_id=house_id))
    _require_write()
    house = House.query.get_or_404(house_id)
    label = f'{house.via}, {house.citta}'
    db.session.delete(house)
    db.session.commit()
    flash(f'Immobile "{label}" eliminato.', 'info')
    return redirect(url_for('aste.index'))


# ── Aggiornamento inline (AJAX) ────────────────────────────────────────────────
@aste_bp.post('/<int:house_id>/inline')
@login_required
def inline_update(house_id: int):
    _require_write()
    house  = House.query.get_or_404(house_id)
    body   = request.get_json(force=True) or {}
    campo  = body.get('field', '')
    valore = body.get('value')

    CAMPI_FLOAT = {'prezzo_base', 'prezzo_aggiudicazione', 'assegno_anticipo',
                   'stima_vendita', 'ricavo_consuntivo', 'mq'}
    CAMPI_TESTO = {'via', 'citta', 'note', 'stato'}
    CAMPI_BOOL  = {'garage'}
    CAMPI_DATA  = {'data_asta'}

    if campo not in CAMPI_FLOAT | CAMPI_TESTO | CAMPI_BOOL | CAMPI_DATA:
        return jsonify({'ok': False, 'error': f'Campo "{campo}" non modificabile'}), 400

    old = str(getattr(house, campo, ''))
    if campo in CAMPI_FLOAT:
        try: setattr(house, campo, float(str(valore).replace(',', '.')))
        except (ValueError, TypeError): return jsonify({'ok': False, 'error': 'Valore non numerico'}), 400
    elif campo in CAMPI_BOOL:
        setattr(house, campo, bool(valore))
    elif campo in CAMPI_DATA:
        try: setattr(house, campo, date.fromisoformat(valore) if valore else None)
        except ValueError: return jsonify({'ok': False, 'error': 'Data non valida'}), 400
    else:
        setattr(house, campo, str(valore or '').strip())

    _audit(house, campo, old, str(getattr(house, campo, '')))
    db.session.commit()

    return jsonify({
        'ok':              True,
        'spese_totali':    _fmt(house.spese_totali),
        'costo_totale':    _fmt(house.costo_totale),
        'guadagno_stimato': _fmt(house.guadagno_stimato),
        'guadagno_reale':   _fmt(house.guadagno_reale),
    })


# ── Spese ─────────────────────────────────────────────────────────────────────
@aste_bp.post('/<int:house_id>/spese')
@login_required
def add_expense(house_id: int):
    _require_write()
    house = House.query.get_or_404(house_id)
    paid_by_id = request.form.get('paid_by_user_id')
    exp = HouseExpense(
        house_id        = house.id,
        categoria       = request.form.get('categoria', 'altro'),
        descrizione     = request.form.get('descrizione', '').strip(),
        importo         = _to_float(request.form.get('importo')),
        data            = _parse_date(request.form.get('data')),
        paid_by_user_id = int(paid_by_id) if paid_by_id and paid_by_id.isdigit() else current_user.id,
        created_by      = current_user.id,
    )
    db.session.add(exp)
    _audit(house, 'spesa', '', f'+{exp.importo}€ {exp.categoria} — {exp.descrizione} (pagato da: {exp.paid_by.name if exp.paid_by else "—"})')
    db.session.commit()

    # Notifica tutti i partecipanti
    _notify_expense(house, exp)

    flash('Spesa aggiunta.', 'success')
    return redirect(url_for('aste.detail', house_id=house_id))


@aste_bp.post('/<int:house_id>/spese/<int:exp_id>/versato')
@login_required
def expense_versato(house_id: int, exp_id: int):
    """Debitore conferma di aver versato la sua quota di questa spesa."""
    _require_write()
    from ..models import HouseExpenseSettlement
    exp = HouseExpense.query.get_or_404(exp_id)
    s   = HouseExpenseSettlement.query.filter_by(
        expense_id=exp_id, user_id=current_user.id).first()
    if not s:
        s = HouseExpenseSettlement(expense_id=exp_id, user_id=current_user.id)
        db.session.add(s)
    s.versato    = not s.versato
    s.versato_il = datetime.utcnow() if s.versato else None
    db.session.commit()
    return jsonify({'ok': True, 'versato': s.versato})


@aste_bp.post('/<int:house_id>/spese/<int:exp_id>/ricevuto')
@login_required
def expense_ricevuto(house_id: int, exp_id: int):
    """Creditore (chi ha pagato) conferma di aver ricevuto le quote."""
    _require_write()
    from ..models import HouseExpenseSettlement
    exp = HouseExpense.query.get_or_404(exp_id)
    if exp.paid_by_user_id != current_user.id:
        return jsonify({'ok': False, 'error': 'Solo chi ha pagato può confermare la ricezione.'}), 403
    s = HouseExpenseSettlement.query.filter_by(
        expense_id=exp_id, user_id=current_user.id).first()
    if not s:
        s = HouseExpenseSettlement(expense_id=exp_id, user_id=current_user.id)
        db.session.add(s)
    s.ricevuto    = not s.ricevuto
    s.ricevuto_il = datetime.utcnow() if s.ricevuto else None
    db.session.commit()
    return jsonify({'ok': True, 'ricevuto': s.ricevuto})


@aste_bp.post('/<int:house_id>/spese/<int:exp_id>/delete')
@login_required
def delete_expense(house_id: int, exp_id: int):
    _require_write()
    exp   = HouseExpense.query.get_or_404(exp_id)
    house = exp.house
    _audit(house, 'spesa', f'{exp.importo}€ {exp.descrizione}', 'ELIMINATA')
    db.session.delete(exp)
    db.session.commit()
    flash('Spesa eliminata.', 'info')
    return redirect(url_for('aste.detail', house_id=house_id))


# ── Scadenze ──────────────────────────────────────────────────────────────────
@aste_bp.post('/<int:house_id>/scadenze')
@login_required
def add_deadline(house_id: int):
    _require_write()
    house = House.query.get_or_404(house_id)
    dl = HouseDeadline(
        house_id      = house.id,
        titolo        = request.form.get('titolo', '').strip(),
        data_scadenza = _parse_date(request.form.get('data_scadenza')),
        giorni_alert  = int(request.form.get('giorni_alert', 7) or 7),
        tipo          = request.form.get('tipo', 'post'),
    )
    db.session.add(dl)
    db.session.commit()
    flash('Scadenza aggiunta.', 'success')
    return redirect(url_for('aste.detail', house_id=house_id))


@aste_bp.post('/<int:house_id>/scadenze/<int:dl_id>/toggle')
@login_required
def toggle_deadline(house_id: int, dl_id: int):
    _require_write()
    dl = HouseDeadline.query.get_or_404(dl_id)
    dl.completata = not dl.completata
    db.session.commit()
    return jsonify({'ok': True, 'completata': dl.completata})


@aste_bp.post('/<int:house_id>/scadenze/<int:dl_id>/delete')
@login_required
def delete_deadline(house_id: int, dl_id: int):
    _require_write()
    dl = HouseDeadline.query.get_or_404(dl_id)
    db.session.delete(dl)
    db.session.commit()
    flash('Scadenza eliminata.', 'info')
    return redirect(url_for('aste.detail', house_id=house_id))


# ── Export Excel ──────────────────────────────────────────────────────────────
@aste_bp.get('/export')
@login_required
def export_excel():
    _require_view()
    houses = House.query.order_by(House.data_asta.desc()).all()
    buf = _build_excel(houses)
    return send_file(io.BytesIO(buf),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'Aste_{date.today().isoformat()}.xlsx')


# ── Cron — alert scadenze ─────────────────────────────────────────────────────
@aste_bp.get('/api/cron/deadlines')
def cron_deadlines():
    secret = current_app.config.get('CRON_SECRET', '')
    if secret and request.args.get('secret') != secret:
        return jsonify({'error': 'unauthorized'}), 401

    today    = date.today()
    notified = 0
    deadlines = HouseDeadline.query.filter_by(completata=False).all()
    urgent    = [d for d in deadlines
                 if d.data_scadenza and d.data_scadenza <= today + timedelta(days=d.giorni_alert)]

    by_house: dict[int, list] = {}
    for dl in urgent:
        by_house.setdefault(dl.house_id, []).append(dl)

    from ..email_service import send_aste_deadline_alert
    for house_id, dls in by_house.items():
        house = House.query.get(house_id)
        if not house: continue
        recipients = list({u.email for u in User.query.filter_by(is_approved=True).all()})
        for email in recipients:
            try:
                send_aste_deadline_alert(email, house, dls)
                notified += 1
            except Exception as exc:
                current_app.logger.warning('Alert scadenza fallito: %s', exc)

    return jsonify({'ok': True, 'notified': notified})


# ── Helpers ───────────────────────────────────────────────────────────────────
def _house_from_form(house):
    if house is None:
        house = House(created_by=current_user.id)
    house.via                   = request.form.get('via', '').strip()
    house.citta                 = request.form.get('citta', '').strip()
    house.mq                    = _to_float(request.form.get('mq')) or None
    house.garage                = bool(request.form.get('garage'))
    house.data_asta             = _parse_date(request.form.get('data_asta'))
    house.prezzo_base           = _to_float(request.form.get('prezzo_base'))
    house.prezzo_aggiudicazione = _to_float(request.form.get('prezzo_aggiudicazione'))
    house.assegno_anticipo      = _to_float(request.form.get('assegno_anticipo'))
    anticipo_uid = request.form.get('anticipo_pagato_da')
    house.anticipo_pagato_da    = int(anticipo_uid) if anticipo_uid and anticipo_uid.isdigit() else None
    house.stima_vendita         = _to_float(request.form.get('stima_vendita'))
    house.ricavo_consuntivo     = _to_float(request.form.get('ricavo_consuntivo'))
    house.stato                 = request.form.get('stato', 'in_asta')
    house.note                  = request.form.get('note', '').strip()
    return house


def _update_from_form(house):
    tracked = ['via','citta','mq','garage','data_asta','prezzo_base',
               'prezzo_aggiudicazione','assegno_anticipo','stima_vendita',
               'ricavo_consuntivo','stato','note']
    old_vals = {f: str(getattr(house, f, '')) for f in tracked}
    _house_from_form(house)
    for campo, old in old_vals.items():
        new = str(getattr(house, campo, ''))
        if old != new:
            _audit(house, campo, old, new)


def _save_participants(house):
    user_ids = [int(x) for x in request.form.getlist('participants') if x.isdigit()]
    HouseParticipant.query.filter_by(house_id=house.id).filter(
        ~HouseParticipant.user_id.in_(user_ids)
    ).delete(synchronize_session=False)
    existing = {p.user_id for p in house.participants.all()}
    for uid in user_ids:
        if uid not in existing:
            db.session.add(HouseParticipant(house_id=house.id, user_id=uid))


def _save_expenses(house):
    cats  = request.form.getlist('exp_cat')
    descs = request.form.getlist('exp_desc')
    imps  = request.form.getlist('exp_imp')
    dates = request.form.getlist('exp_date')
    paids = request.form.getlist('exp_paid_by')
    for cat, desc, imp, d, paid in zip(cats, descs, imps, dates, paids or ['']*len(cats)):
        if not imp: continue
        db.session.add(HouseExpense(
            house_id        = house.id,
            categoria       = cat or 'altro',
            descrizione     = desc.strip(),
            importo         = _to_float(imp),
            data            = _parse_date(d),
            paid_by_user_id = int(paid) if paid and paid.isdigit() else current_user.id,
            created_by      = current_user.id,
        ))


def _save_deadlines(house):
    titoli = request.form.getlist('dl_titolo')
    dates  = request.form.getlist('dl_date')
    giorni = request.form.getlist('dl_giorni')
    tipi   = request.form.getlist('dl_tipo')
    for titolo, d, g, t in zip(titoli, dates, giorni, tipi):
        if not titolo: continue
        db.session.add(HouseDeadline(
            house_id      = house.id,
            titolo        = titolo.strip(),
            data_scadenza = _parse_date(d),
            giorni_alert  = int(g or 7),
            tipo          = t or 'post',
        ))


def _audit(house, campo, prima, dopo):
    db.session.add(HouseAudit(
        house_id     = house.id,
        user_id      = current_user.id,
        campo        = campo,
        valore_prima = prima,
        valore_dopo  = dopo,
    ))


def _notify_expense(house, expense):
    """Notifica tutti i partecipanti della nuova spesa e il loro dare/avere."""
    from ..email_service import send_aste_expense_notification
    parts = house.participants.all()
    if not parts:
        return
    n = house.n_partecipanti
    quota_spesa = expense.importo / n  # quota di questa singola spesa
    payer_name  = expense.paid_by.name if expense.paid_by else 'Sconosciuto'

    for p in parts:
        saldo = house.saldo_utente(p.user_id)
        try:
            send_aste_expense_notification(
                to_email   = p.user.email,
                to_name    = p.user.name,
                house      = house,
                expense    = expense,
                quota_spesa= quota_spesa,
                payer_name = payer_name,
                saldo      = saldo,
            )
        except Exception as exc:
            current_app.logger.warning('Notifica spesa fallita per %s: %s', p.user.email, exc)


def _compute_kpi(houses):
    return {
        'tot_aggiudicazione': sum((h.prezzo_aggiudicazione or 0) for h in houses if h.stato != 'in_asta'),
        'tot_anticipo':       sum((h.assegno_anticipo or 0) for h in houses),
        'tot_spese':          sum(h.spese_totali for h in houses),
        'tot_costo':          sum(h.costo_totale for h in houses if h.stato != 'in_asta'),
        'stima_incassi':      sum((h.stima_vendita or 0) for h in houses if h.stato not in ('in_asta','venduta')),
        'guadagno_stimato':   sum(h.guadagno_stimato for h in houses if h.stato not in ('in_asta','venduta')),
        'ricavo_consuntivo':  sum((h.ricavo_consuntivo or 0) for h in houses if h.stato == 'venduta'),
        'guadagno_reale':     sum(h.guadagno_reale for h in houses if h.stato == 'venduta'),
        'per_stato':          {s: sum(1 for h in houses if h.stato == s) for s in HOUSE_STATI},
        'urgenti':            sum(1 for h in houses for _ in h.scadenze_urgenti),
    }


def _to_float(v) -> float:
    try:
        return float(str(v or '0').replace(',', '.').replace('€', '').strip())
    except (ValueError, TypeError):
        return 0.0


def _parse_date(v):
    if not v: return None
    try: return date.fromisoformat(v)
    except (ValueError, TypeError): return None


def _fmt(v: float) -> str:
    return f'{v:,.2f} €'.replace(',', 'X').replace('.', ',').replace('X', '.')


def _build_excel(houses: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Gestione Aste'
    ws.sheet_view.showGridLines = False

    headers = ['#','Via / Città','mq','G','Data Asta','Stato',
               'Prezzo Base €','Aggiudicazione €','Anticipo €',
               'Spese Tot. €','Costo Tot. €','Stima €','G. Stimato €','Ricavo €','G. Reale €']
    widths = [4,28,6,4,12,14,14,14,12,12,12,12,13,12,12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin   = Side(style='thin', color='D0DCF0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
        c.fill      = PatternFill('solid', fgColor='0D1830')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28

    def m(v): return f'{v:,.2f}'.replace(',','X').replace('.', ',').replace('X','.') if v else '—'
    pc = {'critica':'FFE0E0','alta':'FFF3CD','media':'E8F4FD','bassa':'F4F4F4'}

    for ri, h in enumerate(houses, 2):
        row = [h.id, f'{h.via}, {h.citta}', h.mq or '—',
               '✓' if h.garage else '',
               h.data_asta.strftime('%d/%m/%Y') if h.data_asta else '—',
               HOUSE_STATI.get(h.stato, h.stato),
               m(h.prezzo_base), m(h.prezzo_aggiudicazione), m(h.assegno_anticipo),
               m(h.spese_totali), m(h.costo_totale),
               m(h.stima_vendita), m(h.guadagno_stimato),
               m(h.ricavo_consuntivo), m(h.guadagno_reale)]
        fill = 'F4F8FF' if ri % 2 == 0 else 'FFFFFF'
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.font      = Font(name='Calibri', color='1D2A44', size=10)
            cf = fill
            if ci == 13: cf = 'D5F5E3' if (h.guadagno_stimato or 0) >= 0 else 'FADBD8'
            if ci == 15: cf = 'D5F5E3' if (h.guadagno_reale or 0) >= 0 else 'FADBD8'
            c.fill      = PatternFill('solid', fgColor=cf)
            c.border    = border
            c.alignment = Alignment(vertical='center',
                                    horizontal='right' if ci >= 7 else ('center' if ci in (1,3,4) else 'left'),
                                    indent=1)
        ws.row_dimensions[ri].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
