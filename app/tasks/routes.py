from __future__ import annotations
import io
import json
from datetime import datetime, date, timedelta

from flask import (abort, current_app, flash, jsonify, redirect,
                   render_template, request, send_file, url_for)
from urllib.parse import urlparse
from flask_login import current_user, login_required

from . import tasks_bp
from .. import db
from ..models import (Task, TaskCategory, TaskEvent, TaskHelperRequest,
                      TaskRecipientResponse, TaskAttachment, TASK_STATUS_LABELS, TASK_STATUS_COLORS,
                      TASK_PRIORITY_LABELS, TASK_PRIORITY_COLORS,
                      TASK_STATUS_ORDER)
from ..email_service import send_task_notification, send_task_helper_response
from ..telegram_service import send_task_update_notification, send_task_deadline_alert
from .utils import build_powerapp_task_url


def _require_view():
    if not current_user.is_authenticated or not current_user.can_view('task_manager'):
        abort(403)

def _require_write():
    if not current_user.is_authenticated or not current_user.can_write('task_manager'):
        abort(403)


def _require_admin():
    if not current_user.is_authenticated or not current_user.is_admin:
        abort(403)


def _actor_name(default_email: str = '') -> str:
    if current_user.is_authenticated:
        return current_user.name
    return default_email or 'Sistema'


def _actor_email(default_email: str = '') -> str:
    if current_user.is_authenticated:
        return current_user.email
    return default_email or ''


def _set_completion_fields(task: Task, actor_name: str = '', actor_email: str = '') -> None:
    if task.status == 'completato':
        task.completed_at = datetime.utcnow()
        task.completed_by_name = actor_name or None
        task.completed_by_email = actor_email or None
    elif task.status != 'completato':
        task.completed_at = None
        task.completed_by_name = None
        task.completed_by_email = None




def _safe_telegram_task_update(task: Task, *, action: str, actor_name: str = '', actor_email: str = '', old_status: str = '', new_status: str = '', note: str = '', source: str = '', extra_lines: list[str] | None = None) -> None:
    try:
        send_task_update_notification(
            task,
            action=action,
            actor_name=actor_name,
            actor_email=actor_email,
            old_status=old_status,
            new_status=new_status,
            note=note,
            source=source,
            extra_lines=extra_lines,
        )
    except Exception as exc:
        current_app.logger.warning('Telegram task update failed task_id=%s: %s', task.id, exc)


def _task_edit_extra_lines(task: Task, *, old_title: str = '', old_priority: str = '', old_due_date = None, old_category_name: str = '') -> list[str]:
    lines: list[str] = []
    if old_title and old_title != task.title:
        lines.append(f'Titolo: {old_title} -> {task.title}')
    if old_priority and old_priority != task.priority:
        lines.append(f'Priorità: {old_priority} -> {task.priority}')
    old_due = old_due_date.strftime('%d/%m/%Y') if old_due_date else '—'
    new_due = task.due_date.strftime('%d/%m/%Y') if task.due_date else '—'
    if old_due != new_due:
        lines.append(f'Scadenza: {old_due} -> {new_due}')
    new_cat = task.category.name if task.category else '—'
    if old_category_name and old_category_name != new_cat:
        lines.append(f'Categoria: {old_category_name} -> {new_cat}')
    return lines


def _normalize_task_id(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.replace('#', '').strip()
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _safe_text(value, fallback: str = '') -> str:
    return (value or fallback or '').strip()


def _detect_attachment_provider(url: str) -> str:
    try:
        host = (urlparse(url).hostname or '').lower()
    except Exception:
        host = ''
    if 'sharepoint.com' in host or 'onedrive.live.com' in host or host.endswith('.sharepoint.com'):
        return 'onedrive'
    return 'link'


def _parse_attachment_lines(raw_value: str | None) -> list[dict]:
    items: list[dict] = []
    for line in (raw_value or '').splitlines():
        line = line.strip()
        if not line:
            continue
        label = ''
        url = line
        if '|' in line:
            label, url = [part.strip() for part in line.split('|', 1)]
        if not url.lower().startswith('http://') and not url.lower().startswith('https://'):
            continue
        items.append({
            'label': label[:255] if label else '',
            'url': url,
            'provider': _detect_attachment_provider(url),
        })
    # dedup per url mantenendo ordine
    deduped: list[dict] = []
    seen: set[str] = set()
    for item in items:
        if item['url'] in seen:
            continue
        seen.add(item['url'])
        deduped.append(item)
    return deduped


def _parse_attachment_payload_from_json(body: dict | None) -> list[dict]:
    body = body or {}
    attachments = body.get('attachments')
    parsed: list[dict] = []
    if isinstance(attachments, str):
        try:
            attachments = json.loads(attachments)
        except Exception:
            attachments = None
    if isinstance(attachments, list):
        for item in attachments:
            if not isinstance(item, dict):
                continue
            url = _safe_text(item.get('url')) or _safe_text(item.get('link_url'))
            label = _safe_text(item.get('label')) or _safe_text(item.get('link_label'))
            if not url.lower().startswith('http://') and not url.lower().startswith('https://'):
                continue
            parsed.append({'label': label[:255], 'url': url, 'provider': _detect_attachment_provider(url)})
    single_url = _safe_text(body.get('attachment_url'))
    if single_url.lower().startswith('http://') or single_url.lower().startswith('https://'):
        parsed.append({
            'label': _safe_text(body.get('attachment_label'))[:255],
            'url': single_url,
            'provider': _detect_attachment_provider(single_url),
        })
    if body.get('attachment_links'):
        parsed.extend(_parse_attachment_lines(body.get('attachment_links')))
    deduped: list[dict] = []
    seen: set[str] = set()
    for item in parsed:
        if item['url'] in seen:
            continue
        seen.add(item['url'])
        deduped.append(item)
    return deduped


def _serialize_attachment(att: TaskAttachment) -> dict:
    return {
        'id': att.id,
        'scope': att.attachment_scope,
        'url': att.link_url,
        'label': att.display_label,
        'provider': att.storage_provider or 'link',
        'created_at': att.created_at.isoformat() if att.created_at else None,
        'added_by_name': att.added_by_name or '',
        'added_by_email': att.added_by_email or '',
    }


def _sync_task_opening_attachments(task: Task, attachments: list[dict], *, actor_name: str = '', actor_email: str = '') -> None:
    task.attachments.filter_by(attachment_scope='task').delete(synchronize_session=False)
    for item in attachments:
        db.session.add(TaskAttachment(
            task_id=task.id,
            attachment_scope='task',
            link_url=item['url'],
            link_label=item.get('label') or None,
            storage_provider=item.get('provider') or _detect_attachment_provider(item['url']),
            added_by_name=actor_name or None,
            added_by_email=actor_email or None,
        ))


def _add_response_attachments(task: Task, event: TaskEvent | None, attachments: list[dict], *, actor_name: str = '', actor_email: str = '') -> None:
    if not attachments:
        return
    for item in attachments:
        db.session.add(TaskAttachment(
            task_id=task.id,
            event_id=event.id if event else None,
            attachment_scope='response',
            link_url=item['url'],
            link_label=item.get('label') or None,
            storage_provider=item.get('provider') or _detect_attachment_provider(item['url']),
            added_by_name=actor_name or None,
            added_by_email=actor_email or None,
        ))


def _build_attachment_lines_for_note(attachments: list[dict]) -> list[str]:
    lines: list[str] = []
    for item in attachments or []:
        label = item.get('label') or item.get('url')
        lines.append(f'Allegato: {label} -> {item.get("url")}')
    return lines


def _warmup_task_backend(task_id: int | None = None, token: str | None = None) -> tuple[bool, str]:
    try:
        db.session.execute(db.text('SELECT 1'))
        if task_id is not None and not Task.query.get(task_id):
            return False, f'Task #{task_id} non trovato.'
        if token and not Task.query.filter_by(external_token=token).first():
            return False, 'Task non trovato.'
        return True, 'ready'
    except Exception as exc:
        current_app.logger.warning('Warmup failed task_id=%s token=%s: %s', task_id, token, exc)
        return False, str(exc)


def _get_or_create_recipient_response(task: Task, recipient_email: str) -> TaskRecipientResponse:
    resp = TaskRecipientResponse.query.filter_by(task_id=task.id, recipient_email=recipient_email).first()
    if resp:
        return resp
    resp = TaskRecipientResponse(task_id=task.id, recipient_email=recipient_email, status='da_fare')
    db.session.add(resp)
    db.session.flush()
    return resp


def _send_due_reminders_to_recipients(today: date) -> dict:
    offsets = [
        (2, 'reminder_2d_sent_at', 'reminder_2d_for_due_date', '⏰ Promemoria Task — scadenza tra 2 giorni'),
        (1, 'reminder_1d_sent_at', 'reminder_1d_for_due_date', '⏰ Promemoria Task — scadenza domani'),
        (0, 'reminder_0d_sent_at', 'reminder_0d_for_due_date', '🚨 Promemoria Task — scadenza oggi'),
    ]
    results = {'two_days': 0, 'one_day': 0, 'same_day': 0, 'emails_sent': 0}
    for offset, sent_col, due_col, subject in offsets:
        target_date = today + timedelta(days=offset)
        tasks = Task.query.filter(
            Task.status.notin_(['completato', 'annullato']),
            Task.due_date == target_date,
        ).order_by(Task.priority.desc(), Task.id.asc()).all()
        grouped: dict[str, list[Task]] = {}
        touched: list[tuple[TaskRecipientResponse, date, str]] = []
        for task in tasks:
            for recipient in task._recipients():
                resp = _get_or_create_recipient_response(task, recipient)
                if getattr(resp, due_col) == task.due_date:
                    continue
                grouped.setdefault(recipient, []).append(task)
                touched.append((resp, task.due_date, sent_col))
        for recipient, recipient_tasks in grouped.items():
            try:
                ok = send_task_notification(
                    to_email=recipient,
                    subject=subject,
                    tasks=recipient_tasks,
                    template_type='reminder',
                    extra={'reminder_days_before_due': offset},
                )
            except Exception as exc:
                ok = False
                current_app.logger.warning('Reminder send failed recipient=%s offset=%s: %s', recipient, offset, exc)
            if not ok:
                continue
            results['emails_sent'] += 1
            key = {2: 'two_days', 1: 'one_day', 0: 'same_day'}[offset]
            results[key] += len(recipient_tasks)
            now = datetime.utcnow()
            for resp, due_date, flag_name in [t for t in touched if t[0].recipient_email == recipient]:
                setattr(resp, flag_name, now)
                setattr(resp, flag_name.replace('_sent_at', '_for_due_date'), due_date)
        db.session.commit()
    return results

def _notify_task_creator_update(task: Task, actor_name: str, actor_email: str, new_status: str, note: str, attachments: list[dict] | None = None) -> None:
    creator_email = (task.created_by_email or '').strip().lower()
    if not creator_email:
        return
    if actor_email and creator_email == actor_email.strip().lower():
        return
    try:
        send_task_notification(
            to_email=creator_email,
            subject=f'💬 Aggiornamento task [TASK-{task.id}] {task.title}',
            tasks=[task],
            template_type='thread_update',
            extra={
                'actor_name': actor_name,
                'actor_email': actor_email,
                'new_status': new_status,
                'note': note,
                'attachments': attachments or [],
            },
        )
    except Exception as exc:
        current_app.logger.warning('Invio aggiornamento creator fallito task_id=%s: %s', task.id, exc)


@tasks_bp.get('/open/<int:task_id>')
def open_in_powerapp(task_id: int):
    task = Task.query.get_or_404(task_id)
    deep_link = build_powerapp_task_url(task.id, task.external_token)
    if not deep_link:
        current_app.logger.warning('POWERAPP_URL non configurato per task_id=%s', task.id)
        return render_template('tasks/powerapp_redirect_missing.html', task=task), 503
    return render_template('tasks/powerapp_warmup.html', task=task, deep_link=deep_link)


@tasks_bp.get('/api/warmup')
def api_warmup():
    task_id = _normalize_task_id(request.args.get('task_id'))
    token = (request.args.get('task_token') or '').strip()
    ok, reason = _warmup_task_backend(task_id=task_id, token=token or None)
    return jsonify({'ok': ok, 'reason': reason, 'task_id': task_id, 'task_token': token})


# ── Board principale ──────────────────────────────────────────────────────────
@tasks_bp.get('')
@login_required
def board():
    _require_view()
    view = request.args.get('view', 'kanban')
    cats = TaskCategory.query.order_by(TaskCategory.position, TaskCategory.name).all()

    f_cat = [c for c in request.args.getlist('cat') if c]
    f_priority = [p for p in request.args.getlist('priority') if p]
    f_status = [s for s in request.args.getlist('status') if s]
    f_search   = request.args.get('search', '').strip()
    f_due_from = request.args.get('due_from', '')
    f_due_to   = request.args.get('due_to', '')

    q = Task.query
    if f_cat:
        q = q.filter(Task.category_id.in_([int(c) for c in f_cat if c.isdigit()]))
    if f_priority:
        q = q.filter(Task.priority.in_(f_priority))
    if f_status:
        q = q.filter(Task.status.in_(f_status))
    if f_search:
        like = f'%{f_search}%'
        q = q.filter(db.or_(Task.title.ilike(like), Task.description.ilike(like)))
    if f_due_from:
        try: q = q.filter(Task.due_date >= date.fromisoformat(f_due_from))
        except ValueError: pass
    if f_due_to:
        try: q = q.filter(Task.due_date <= date.fromisoformat(f_due_to))
        except ValueError: pass

    tasks = q.order_by(
        db.case({'critica': 0, 'alta': 1, 'media': 2, 'bassa': 3}, value=Task.priority),
        Task.due_date.asc().nullslast(),
        Task.created_at.desc()
    ).all()

    today = date.today()
    week_ago = datetime.utcnow() - timedelta(days=7)
    kpi = {
        'da_fare':    sum(1 for t in tasks if t.status == 'da_fare'),
        'in_corso':   sum(1 for t in tasks if t.status == 'in_corso'),
        'in_attesa':  sum(1 for t in tasks if t.status == 'in_attesa'),
        'scaduti':    sum(1 for t in tasks if t.is_overdue),
        'completati_settimana': Task.query.filter(
            Task.status == 'completato', Task.updated_at >= week_ago).count(),
    }

    kanban = {s: [] for s in TASK_STATUS_ORDER}
    for t in tasks:
        kanban[t.status].append(t)

    return render_template('tasks/board.html',
                           view=view, tasks=tasks, kanban=kanban,
                           categories=cats, kpi=kpi,
                           status_labels=TASK_STATUS_LABELS,
                           status_colors=TASK_STATUS_COLORS,
                           priority_labels=TASK_PRIORITY_LABELS,
                           priority_colors=TASK_PRIORITY_COLORS,
                           status_order=TASK_STATUS_ORDER,
                           filters=dict(cat=f_cat, priority=f_priority,
                                        status=f_status, search=f_search,
                                        due_from=f_due_from, due_to=f_due_to))


@tasks_bp.get('/kpi')
@login_required
def kpi_dashboard():
    _require_admin()

    tasks = Task.query.order_by(Task.created_at.desc()).all()
    responses = TaskRecipientResponse.query.order_by(TaskRecipientResponse.updated_at.desc()).all()
    overdue_open = [t for t in tasks if t.is_overdue]
    completed = [t for t in tasks if t.status == 'completato']
    completed_on_time = [t for t in completed if t.completed_in_time is True]
    completed_late = [t for t in completed if t.completed_in_time is False]

    by_person: dict[str, dict] = {}
    for r in responses:
        key = (r.recipient_email or '').strip().lower() or 'sconosciuto'
        bucket = by_person.setdefault(key, {
            'email': key,
            'label': r.replied_by or key,
            'responses': 0,
            'last_response_at': None,
            'late_count': 0,
            'on_time_count': 0,
        })
        bucket['responses'] += 1
        if not bucket['last_response_at'] or (r.updated_at and r.updated_at > bucket['last_response_at']):
            bucket['last_response_at'] = r.updated_at
        if r.task and r.task.status == 'completato':
            if r.task.completed_in_time is True:
                bucket['on_time_count'] += 1
            elif r.task.completed_in_time is False:
                bucket['late_count'] += 1

    responders = sorted(by_person.values(), key=lambda x: (-x['responses'], x['email']))[:10]
    top_late = sorted(by_person.values(), key=lambda x: (-x['late_count'], -x['responses'], x['email']))[:10]
    top_on_time = sorted(by_person.values(), key=lambda x: (-x['on_time_count'], -x['responses'], x['email']))[:10]

    recent_events = TaskEvent.query.filter(TaskEvent.event_type.in_(['external_reply', 'status_changed']))\
        .order_by(TaskEvent.created_at.desc()).limit(30).all()

    status_counts = []
    for key in TASK_STATUS_ORDER:
        count = sum(1 for t in tasks if t.status == key)
        status_counts.append({
            'key': key,
            'label': TASK_STATUS_LABELS.get(key, key),
            'count': count,
            'color': TASK_STATUS_COLORS.get(key, '#4a5c78'),
        })

    total_tasks = len(tasks) or 1
    return render_template(
        'tasks/kpi.html',
        cards={
            'total_tasks': len(tasks),
            'responses_total': len(responses),
            'completed_on_time': len(completed_on_time),
            'completed_late': len(completed_late),
            'overdue_open': len(overdue_open),
            'response_rate': round((len({r.task_id for r in responses}) / total_tasks) * 100),
        },
        status_counts=status_counts,
        responders=responders,
        top_late=top_late,
        top_on_time=top_on_time,
        overdue_open=overdue_open[:20],
        recent_events=recent_events,
        status_labels=TASK_STATUS_LABELS,
        status_colors=TASK_STATUS_COLORS,
    )


# ── Crea task ─────────────────────────────────────────────────────────────────
@tasks_bp.get('/new')
@login_required
def new_task():
    _require_write()
    cats = TaskCategory.query.order_by(TaskCategory.position, TaskCategory.name).all()
    return render_template('tasks/form.html', task=None, categories=cats,
                           priority_labels=TASK_PRIORITY_LABELS,
                           status_labels=TASK_STATUS_LABELS)


@tasks_bp.post('/new')
@login_required
def create_task():
    _require_write()
    title       = request.form.get('title', '').strip()
    description = request.form.get('description', '').strip()
    category_id = request.form.get('category_id', '')
    priority    = request.form.get('priority', 'media')
    status      = request.form.get('status', 'da_fare')
    start_date  = _parse_date(request.form.get('start_date'))
    due_date    = _parse_date(request.form.get('due_date'))
    opening_attachments = _parse_attachment_lines(request.form.get('attachment_links'))

    errors = []
    if not title:       errors.append('Il titolo è obbligatorio.')
    if not category_id: errors.append('Seleziona una categoria.')

    if errors:
        for e in errors: flash(e, 'error')
        cats = TaskCategory.query.order_by(TaskCategory.position).all()
        return render_template('tasks/form.html', task=None, categories=cats,
                               priority_labels=TASK_PRIORITY_LABELS,
                               status_labels=TASK_STATUS_LABELS), 422

    task = Task(title=title, description=description,
                category_id=int(category_id), priority=priority,
                status=status, start_date=start_date, due_date=due_date,
                created_by_user_id=current_user.id if current_user.is_authenticated else None,
                created_by_name=current_user.name if current_user.is_authenticated else None,
                created_by_email=current_user.email if current_user.is_authenticated else None)
    db.session.add(task)
    db.session.flush()

    ev = TaskEvent(task_id=task.id, event_type='created',
                   new_status=status, note='Task creato.',
                   actor_name=_actor_name(), actor_email=_actor_email())
    db.session.add(ev)
    db.session.flush()
    _sync_task_opening_attachments(task, opening_attachments,
                                   actor_name=_actor_name(), actor_email=_actor_email())
    db.session.commit()

    _notify_category(task)
    _safe_telegram_task_update(
        task,
        action='Creato',
        actor_name=_actor_name(),
        actor_email=_actor_email(),
        new_status=task.status,
        source='web',
    )

    flash(f'Task "{title}" creato.', 'success')
    return redirect(url_for('tasks.board'))


# ── Dettaglio task ────────────────────────────────────────────────────────────
@tasks_bp.get('/<int:task_id>')
@login_required
def detail(task_id: int):
    _require_view()
    task = Task.query.get_or_404(task_id)
    cats = TaskCategory.query.order_by(TaskCategory.position).all()
    return render_template('tasks/detail.html', task=task, categories=cats,
                           status_labels=TASK_STATUS_LABELS,
                           status_colors=TASK_STATUS_COLORS,
                           priority_labels=TASK_PRIORITY_LABELS,
                           priority_colors=TASK_PRIORITY_COLORS)


# ── Modifica task ─────────────────────────────────────────────────────────────
@tasks_bp.get('/<int:task_id>/edit')
@login_required
def edit_task(task_id: int):
    _require_write()
    task = Task.query.get_or_404(task_id)
    cats = TaskCategory.query.order_by(TaskCategory.position).all()
    return render_template('tasks/form.html', task=task, categories=cats,
                           priority_labels=TASK_PRIORITY_LABELS,
                           status_labels=TASK_STATUS_LABELS)


@tasks_bp.post('/<int:task_id>/edit')
@login_required
def update_task(task_id: int):
    _require_write()
    task       = Task.query.get_or_404(task_id)
    title      = request.form.get('title', '').strip()
    old_status = task.status
    old_title = task.title
    old_priority = task.priority
    old_due_date = task.due_date
    old_category_name = task.category.name if task.category else '—'
    new_status = request.form.get('status', task.status)
    opening_attachments = _parse_attachment_lines(request.form.get('attachment_links'))

    if not title:
        flash('Il titolo è obbligatorio.', 'error')
        cats = TaskCategory.query.order_by(TaskCategory.position).all()
        return render_template('tasks/form.html', task=task, categories=cats,
                               priority_labels=TASK_PRIORITY_LABELS,
                               status_labels=TASK_STATUS_LABELS), 422

    task.title       = title
    task.description = request.form.get('description', '').strip()
    task.category_id = int(request.form.get('category_id', task.category_id))
    task.priority    = request.form.get('priority', task.priority)
    task.status      = new_status
    task.start_date  = _parse_date(request.form.get('start_date'))
    task.due_date    = _parse_date(request.form.get('due_date'))

    if old_status != new_status:
        ev = TaskEvent(task_id=task.id, event_type='status_changed',
                       old_status=old_status, new_status=new_status,
                       actor_name=_actor_name(), actor_email=_actor_email())
        db.session.add(ev)
    _set_completion_fields(task, _actor_name(), _actor_email())
    db.session.flush()
    _sync_task_opening_attachments(task, opening_attachments,
                                   actor_name=_actor_name(), actor_email=_actor_email())

    db.session.commit()
    _safe_telegram_task_update(
        task,
        action='Aggiornato da web',
        actor_name=_actor_name(),
        actor_email=_actor_email(),
        old_status=old_status,
        new_status=task.status,
        source='web',
        extra_lines=_task_edit_extra_lines(task, old_title=old_title, old_priority=old_priority, old_due_date=old_due_date, old_category_name=old_category_name),
    )
    flash('Task aggiornato.', 'success')
    return redirect(url_for('tasks.board'))


# ── Elimina task ──────────────────────────────────────────────────────────────
@tasks_bp.post('/<int:task_id>/delete')
@login_required
def delete_task(task_id: int):
    _require_write()
    task  = Task.query.get_or_404(task_id)
    title = task.title
    db.session.delete(task)
    db.session.commit()
    flash(f'Task "{title}" eliminato.', 'info')
    return redirect(url_for('tasks.board'))


# ── Cambio stato rapido (Kanban AJAX) ─────────────────────────────────────────
@tasks_bp.post('/<int:task_id>/status')
@login_required
def update_status(task_id: int):
    _require_write()
    task       = Task.query.get_or_404(task_id)
    old_status = task.status
    new_status = (request.get_json(force=True) or {}).get('status', '')

    if new_status not in TASK_STATUS_ORDER:
        return jsonify({'ok': False, 'error': 'Stato non valido'}), 400

    task.status = new_status
    _set_completion_fields(task, _actor_name(), _actor_email())
    ev = TaskEvent(task_id=task.id, event_type='status_changed',
                   old_status=old_status, new_status=new_status,
                   actor_name=_actor_name(), actor_email=_actor_email())
    db.session.add(ev)
    db.session.commit()
    _safe_telegram_task_update(
        task,
        action='Cambio stato rapido',
        actor_name=_actor_name(),
        actor_email=_actor_email(),
        old_status=old_status,
        new_status=new_status,
        source='kanban',
    )
    return jsonify({'ok': True, 'status': new_status,
                    'label': TASK_STATUS_LABELS[new_status],
                    'color': TASK_STATUS_COLORS[new_status]})


# ── Export Excel ──────────────────────────────────────────────────────────────
@tasks_bp.get('/export')
@login_required
def export_excel():
    _require_view()
    tasks     = Task.query.order_by(Task.status, Task.due_date).all()
    xlsx_bytes = _build_excel(tasks)
    return send_file(io.BytesIO(xlsx_bytes),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'Task_{date.today().isoformat()}.xlsx')


# ── Form pubblico risposta (no login) ─────────────────────────────────────────
@tasks_bp.get('/respond/<token>')
def respond_form(token: str):
    task = Task.query.filter_by(external_token=token).first_or_404()
    if task.status in ('completato', 'annullato'):
        return render_template('tasks/respond_closed.html', task=task,
                               status_labels=TASK_STATUS_LABELS)
    return render_template('tasks/respond.html', task=task,
                           status_labels=TASK_STATUS_LABELS,
                           status_colors=TASK_STATUS_COLORS,
                           priority_labels=TASK_PRIORITY_LABELS,
                           priority_colors=TASK_PRIORITY_COLORS)


@tasks_bp.post('/respond/<token>')
def respond_submit(token: str):
    task = Task.query.filter_by(external_token=token).first_or_404()
    if task.status in ('completato', 'annullato'):
        return render_template('tasks/respond_closed.html', task=task,
                               status_labels=TASK_STATUS_LABELS)

    new_status  = request.form.get('status', '').strip()
    note        = request.form.get('note', '').strip()
    replied_by  = request.form.get('replied_by', '').strip()
    response_attachments = _parse_attachment_lines(request.form.get('attachment_links'))
    actor_email = ''
    actor_name = replied_by or 'Utente esterno'

    if new_status not in TASK_STATUS_ORDER:
        flash('Seleziona uno stato valido.', 'error')
        return redirect(url_for('tasks.respond_form', token=token))

    old_status  = task.status
    task.status = new_status
    _set_completion_fields(task, actor_name, actor_email)
    event_note = note
    attachment_lines = _build_attachment_lines_for_note(response_attachments)
    if attachment_lines:
        event_note = (event_note + '\n' if event_note else '') + '\n'.join(attachment_lines)
    ev = TaskEvent(task_id=task.id, event_type='external_reply',
                   old_status=old_status, new_status=new_status,
                   note=event_note,
                   source='form', actor_name=actor_name, actor_email=actor_email)
    db.session.add(ev)
    db.session.flush()
    _add_response_attachments(task, ev, response_attachments,
                              actor_name=actor_name, actor_email=actor_email)
    db.session.commit()

    _notify_task_creator_update(task, actor_name, actor_email, new_status, event_note, response_attachments)
    _safe_telegram_task_update(
        task,
        action='Aggiornamento da form pubblico',
        actor_name=actor_name,
        actor_email=actor_email,
        old_status=old_status,
        new_status=new_status,
        note=event_note,
        source='form',
    )

    try:
        send_task_notification(
            to_email=current_app.config['ADMIN_EMAIL'],
            subject=f'↩ Risposta su [TASK-{task.id}] {task.title}',
            tasks=[task],
            template_type='external_reply',
            extra={
                'replied_by': actor_name,
                'note': event_note,
                'new_status': new_status,
                'new_task_status': task.status,
                'aggregate_changed': task.status != old_status,
                'attachments': response_attachments,
            },
        )
    except Exception as exc:
        current_app.logger.warning('Notifica risposta fallita: %s', exc)

    return render_template('tasks/respond_thanks.html', task=task,
                           status_labels=TASK_STATUS_LABELS,
                           new_status=new_status)


# ── API webhook Power Automate ────────────────────────────────────────────────
@tasks_bp.post('/api/<token>/update')
def api_update(token: str):
    api_key = current_app.config.get('TASKS_API_KEY', '')
    if api_key and request.headers.get('X-API-Key') != api_key:
        return jsonify({'error': 'unauthorized'}), 401

    task = Task.query.filter_by(external_token=token).first()
    if not task:
        return jsonify({'error': 'task not found'}), 404

    body = request.get_json(silent=True) or {}
    if not body and request.form:
        body = request.form.to_dict(flat=True)
    new_status = _safe_text(body.get('status')).lower()
    note = _safe_text(body.get('note'))
    replied_by = _safe_text(body.get('replied_by'))
    email = _safe_text(body.get('email') or replied_by).lower()
    response_attachments = _parse_attachment_payload_from_json(body)

    if new_status not in TASK_STATUS_ORDER:
        return jsonify({'error': f'status non valido: {new_status}'}), 400

    if email:
        result = _upsert_recipient_response(task, email, new_status, note, replied_by)
        response_event = result.get('response_event')
        if response_attachments:
            _add_response_attachments(task, response_event, response_attachments,
                                      actor_name=replied_by or email, actor_email=email)
        db.session.commit()
    else:
        old_status = task.status
        task.status = new_status
        _set_completion_fields(task, replied_by or email, email)
        event_note = note
        attachment_lines = _build_attachment_lines_for_note(response_attachments)
        if attachment_lines:
            event_note = (event_note + '\n' if event_note else '') + '\n'.join(attachment_lines)
        ev = TaskEvent(task_id=task.id, event_type='external_reply',
                       old_status=old_status, new_status=new_status,
                       note=event_note,
                       source='powerautomate',
                       actor_name=replied_by or email,
                       actor_email=email)
        db.session.add(ev)
        db.session.flush()
        _add_response_attachments(task, ev, response_attachments,
                                  actor_name=replied_by or email, actor_email=email)
        db.session.commit()
        result = {
            'new_task_status': task.status,
            'aggregate_changed': task.status != old_status,
            'progress_percent': task.progress_percent,
            'response_event': ev,
        }

    event_note = note
    attachment_lines = _build_attachment_lines_for_note(response_attachments)
    if attachment_lines:
        event_note = (event_note + '\n' if event_note else '') + '\n'.join(attachment_lines)

    try:
        send_task_notification(
            to_email=current_app.config['ADMIN_EMAIL'],
            subject=f'↩ Risposta [TASK-{task.id}] via Power Automate',
            tasks=[task],
            template_type='external_reply',
            extra={
                'replied_by': replied_by or email,
                'note': event_note,
                'new_status': new_status,
                'new_task_status': result.get('new_task_status', task.status),
                'aggregate_changed': result.get('aggregate_changed', False),
                'attachments': response_attachments,
            },
        )
    except Exception:
        pass

    _notify_task_creator_update(task, replied_by or email, email, new_status, event_note, response_attachments)
    _safe_telegram_task_update(
        task,
        action='Aggiornamento Power Automate (token)',
        actor_name=replied_by or email,
        actor_email=email,
        old_status=task.status if not result.get('aggregate_changed') else '',
        new_status=result.get('new_task_status', task.status),
        note=event_note,
        source='powerautomate',
        extra_lines=[f'Stato destinatario: {new_status}', f'Avanzamento: {result.get("progress_percent")}% ' if result.get('progress_percent') is not None else '', *_build_attachment_lines_for_note(response_attachments)],
    )

    return jsonify({
        'ok': True,
        'task_id': task.id,
        'recipient_status': new_status,
        'task_status': task.status,
        'task_status_label': TASK_STATUS_LABELS.get(task.status, task.status),
        'aggregate_changed': result.get('aggregate_changed', False),
        'progress_percent': result.get('progress_percent'),
        'attachments': [_serialize_attachment(att) for att in task.attachments.filter_by(attachment_scope='response').all()],
    })


# ── API helper email (Power Automate → invia task a chi ha chiesto) ────────────
@tasks_bp.post('/api/helper')
def api_helper():
    """
    Power Automate chiama questo endpoint quando arriva una mail
    all'indirizzo helper. Rate limit: 1 richiesta/giorno per email.
    """
    api_key = current_app.config.get('TASKS_API_KEY', '')
    if api_key and request.headers.get('X-API-Key') != api_key:
        return jsonify({'error': 'unauthorized'}), 401

    body           = request.get_json(force=True) or {}
    requester_email = (body.get('email') or '').strip().lower()

    if not requester_email or '@' not in requester_email:
        return jsonify({'error': 'email mancante o non valida'}), 400

    # Rate limit: 1 al giorno per email
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    existing = TaskHelperRequest.query.filter(
        TaskHelperRequest.requester_email == requester_email,
        TaskHelperRequest.requested_at >= today_start,
    ).first()

    if existing:
        return jsonify({
            'ok': False,
            'error': 'rate_limited',
            'message': 'Hai già ricevuto il riepilogo oggi. Riprova domani.'
        }), 429

    # Cerca i task associati a questa email (come destinatario categoria)
    all_categories = TaskCategory.query.filter_by(notify_on_create=True).all()
    relevant_cats  = [
        c for c in all_categories
        if requester_email in [(r or '').strip().lower() for r in c.get_recipients()]
    ]
    cat_ids        = [c.id for c in relevant_cats]

    if cat_ids:
        user_tasks = Task.query.filter(
            Task.category_id.in_(cat_ids),
            Task.status.notin_(['completato', 'annullato']),
        ).order_by(Task.due_date.asc().nullslast()).all()
    else:
        user_tasks = []

    # Registra la richiesta
    record = TaskHelperRequest(requester_email=requester_email)
    db.session.add(record)
    db.session.commit()

    # Invia email con riepilogo
    try:
        send_task_helper_response(
            to_email=requester_email,
            tasks=user_tasks,
        )
    except Exception as exc:
        current_app.logger.error('Helper email fallita: %s', exc)
        return jsonify({'ok': False, 'error': 'invio email fallito'}), 500

    return jsonify({
        'ok': True,
        'tasks_found': len(user_tasks),
        'email_sent_to': requester_email,
    })


# ── Cron reminders ────────────────────────────────────────────────────────────
@tasks_bp.get('/api/cron/reminders')
def cron_reminders():
    secret = current_app.config.get('CRON_SECRET', '')
    if secret and request.args.get('secret') != secret:
        return jsonify({'error': 'unauthorized'}), 401

    today = date.today()
    tomorrow = today + timedelta(days=1)

    open_tasks = Task.query.filter(
        Task.status.notin_(['completato', 'annullato']),
        Task.due_date.isnot(None),
    ).all()

    overdue = [t for t in open_tasks if t.due_date < today]
    due_today = [t for t in open_tasks if t.due_date == today]
    due_tomorrow = [t for t in open_tasks if t.due_date == tomorrow]
    urgent = overdue + due_today + due_tomorrow

    admin_notified = 0
    if urgent:
        try:
            send_task_notification(
                to_email=current_app.config['ADMIN_EMAIL'],
                subject=f'⚠ LogiMetric — {len(urgent)} task urgenti',
                tasks=urgent,
                template_type='reminder',
                extra={'reminder_days_before_due': 0},
            )
            admin_notified = len(urgent)
        except Exception:
            pass

        try:
            send_task_deadline_alert(
                overdue_tasks=overdue,
                due_today_tasks=due_today,
                due_tomorrow_tasks=due_tomorrow,
            )
        except Exception as exc:
            current_app.logger.warning('Telegram deadline alert failed: %s', exc)

    reminder_stats = _send_due_reminders_to_recipients(today)

    if today.weekday() == 0:  # lunedì
        all_open = Task.query.filter(Task.status.notin_(['completato', 'annullato'])).all()
        try:
            send_task_notification(
                to_email=current_app.config['ADMIN_EMAIL'],
                subject='📋 LogiMetric — Report settimanale task aperti',
                tasks=all_open,
                template_type='weekly_report',
            )
        except Exception:
            pass

    return jsonify({
        'ok': True,
        'admin_notified': admin_notified,
        'advance_reminders': reminder_stats,
        'date': today.isoformat(),
    })




@tasks_bp.post('/api/telegram/test')
@login_required
def api_telegram_test():
    _require_admin()
    sample = Task.query.order_by(Task.id.desc()).first()
    if sample:
        ok = send_task_update_notification(
            sample,
            action='Test Telegram',
            actor_name=_actor_name(),
            actor_email=_actor_email(),
            new_status=sample.status,
            source='manual-test',
        )
    else:
        ok = send_task_deadline_alert(overdue_tasks=[], due_today_tasks=[], due_tomorrow_tasks=[])
        if not ok:
            from ..telegram_service import send_telegram_message
            ok = send_telegram_message('✅ LogiMetric — test Telegram eseguito correttamente.')
    return jsonify({'ok': bool(ok)})

# ── Builder helpers ───────────────────────────────────────────────────────────
def _resolve_email_in_pool(incoming: str, pool: list[str]) -> str:
    """
    Risolve l'email incoming all'email canonica più simile nel pool.
    Usata sia in tasks-for-user che in _resolve_recipient_email.

    Strategie in ordine:
    1. Match esatto
    2. Match per username (parte prima di @)
    3. Match per User model (se l'utente è registrato in LogiMetric)
    4. Fallback: restituisce incoming invariata
    """
    incoming = incoming.strip().lower()
    if not pool:
        return incoming

    # 1. Esatto
    if incoming in pool:
        return incoming

    # 2. Username
    incoming_user = incoming.split('@')[0]
    for candidate in pool:
        if candidate.split('@')[0] == incoming_user:
            return candidate

    # 3. User model — cerca utente con questa email e usa la sua email
    #    (copre il caso: MS account diverso dall'email LogiMetric)
    from ..models import User
    user = User.query.filter(
        db.func.lower(User.email) == incoming
    ).first()
    if user:
        user_email = user.email.strip().lower()
        if user_email in pool:
            return user_email
        # Prova username dell'email utente vs pool
        user_name = user_email.split('@')[0]
        for candidate in pool:
            if candidate.split('@')[0] == user_name:
                return candidate

    return incoming


def _resolve_recipient_email(task: Task, incoming_email: str) -> str:
    """
    Risolve l'email ricevuta dalla Power App all'email reale del destinatario.

    Strategie:
    1-3. _resolve_email_in_pool (esatto, username, User model)
    4.   Unico destinatario senza risposta → assegna a lui
    5.   Fallback: usa incoming invariata (loggato come warning)
    """
    from ..models import TaskRecipientResponse

    incoming = incoming_email.strip().lower()
    recipients = [r.strip().lower() for r in task._recipients()]

    if not recipients:
        return incoming

    # Strategie 1-3
    resolved = _resolve_email_in_pool(incoming, recipients)
    if resolved in recipients:
        if resolved != incoming:
            current_app.logger.info(
                'Fuzzy match (pool): %s → %s (task #%d)', incoming, resolved, task.id)
        return resolved

    # Strategia 4: unico senza risposta
    already_responded = {
        r.recipient_email for r in
        TaskRecipientResponse.query.filter_by(task_id=task.id).all()
    }
    without_response = [r for r in recipients if r not in already_responded]
    if len(without_response) == 1:
        current_app.logger.info(
            'Fuzzy match (esclusione): %s → %s (task #%d)',
            incoming, without_response[0], task.id)
        return without_response[0]

    # Nessun match → usa incoming e logga per debug
    current_app.logger.warning(
        'Email non risolta: incoming=%s, recipients=%s, task=#%d',
        incoming, recipients, task.id)
    return incoming


def _upsert_recipient_response(task: Task, email: str, new_status: str,
                               note: str = '', replied_by: str = '') -> dict:
    """
    Salva/aggiorna la risposta individuale di un destinatario.
    Usa _resolve_recipient_email per gestire mismatch di email Microsoft vs Gmail.
    """
    from ..models import TaskRecipientResponse

    # Risolvi l'email al destinatario corretto
    resolved_email = _resolve_recipient_email(task, email)

    # Upsert risposta individuale
    resp = TaskRecipientResponse.query.filter_by(
        task_id=task.id, recipient_email=resolved_email).first()
    if resp:
        resp.status     = new_status
        resp.note       = note
        resp.replied_by = replied_by or email  # salva email originale come riferimento
    else:
        resp = TaskRecipientResponse(
            task_id=task.id,
            recipient_email=resolved_email,
            status=new_status,
            note=note,
            replied_by=replied_by or email,
        )
        db.session.add(resp)

    db.session.flush()

    # Ricalcola stato aggregato
    old_task_status = task.status
    agg_status      = task.compute_aggregate_status()
    changed         = agg_status != old_task_status

    if changed:
        task.status = agg_status
        _set_completion_fields(task, replied_by or email, resolved_email)
        ev = TaskEvent(
            task_id=task.id,
            event_type='status_changed',
            old_status=old_task_status,
            new_status=agg_status,
            note=f'Aggiornamento automatico: tutti i destinatari hanno raggiunto '
                 f'"{TASK_STATUS_LABELS.get(agg_status, agg_status)}"',
            source='aggregate',
            actor_name=replied_by or email,
            actor_email=resolved_email,
        )
        db.session.add(ev)

    # Log risposta individuale (include email originale per debug)
    debug_note = note or ''
    if resolved_email != email.strip().lower():
        debug_note = (debug_note + '\n' if debug_note else '') + f'Email originale: {email}'
    ev_individual = TaskEvent(
        task_id=task.id,
        event_type='external_reply',
        old_status=old_task_status,
        new_status=new_status,
        note=debug_note,
        source='powerapp',
        actor_name=replied_by or email,
        actor_email=resolved_email,
    )
    db.session.add(ev_individual)
    db.session.flush()

    return {
        'old_task_status':  old_task_status,
        'new_task_status':  task.status,
        'aggregate_changed': changed,
        'recipient_status': new_status,
        'resolved_email':   resolved_email,
        'progress_percent': task.progress_percent,
        'response_event':   ev_individual,
    }


def _task_to_dict(t: Task) -> dict:
    """Serializza un Task in dizionario JSON per la Power App."""
    task_attachments = [_serialize_attachment(att) for att in t.attachments.filter_by(attachment_scope='task').all()]
    response_attachments = [_serialize_attachment(att) for att in t.attachments.filter_by(attachment_scope='response').all()]
    return {
        'id':               t.id,
        'title':            t.title,
        'description':      t.description or '',
        'category':         t.category.name if t.category else '',
        'priority':         t.priority,
        'priority_label':   TASK_PRIORITY_LABELS.get(t.priority, t.priority),
        'status':           t.status,
        'status_label':     TASK_STATUS_LABELS.get(t.status, t.status),
        'start_date':       t.start_date.isoformat() if t.start_date else '',
        'due_date':         t.due_date.isoformat() if t.due_date else '',
        'due_date_display': t.due_date.strftime('%d/%m/%Y') if t.due_date else '—',
        'is_overdue':       t.is_overdue,
        'token':            t.external_token,
        'progress_percent': t.progress_percent,   # None se 1 solo destinatario
        'attachments':      task_attachments,
        'response_attachments': response_attachments,
    }


def _parse_date(value):
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def _notify_category(task: Task):
    cat = task.category
    if not cat or not cat.notify_on_create:
        return
    recipients = cat.get_recipients()
    if not recipients:
        return
    try:
        send_task_notification(
            to_email=recipients,
            subject=f'[TASK-{task.id}] {task.title}',
            tasks=[task],
            template_type='new_task',
        )
    except Exception as exc:
        current_app.logger.warning('Notifica categoria fallita: %s', exc)


def _build_excel(tasks: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Task'
    ws.sheet_view.showGridLines = False

    widths = [6, 30, 14, 12, 12, 14, 14, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    headers = ['#', 'Titolo', 'Categoria', 'Priorità', 'Stato', 'Inizio', 'Scadenza', 'Creato il']
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font      = Font(bold=True, color='FFFFFF', name='Calibri')
        c.fill      = PatternFill('solid', fgColor='0D1830')
        c.alignment = Alignment(horizontal='center', vertical='center', indent=1)
    ws.row_dimensions[1].height = 22

    prio_colors = {'critica': 'FFE0E0', 'alta': 'FFF3CD', 'media': 'E8F4FD', 'bassa': 'F4F4F4'}
    thin   = Side(style='thin', color='D0DCF0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_n, task in enumerate(tasks, 2):
        row_data = [
            task.id, task.title,
            task.category.name if task.category else '',
            TASK_PRIORITY_LABELS.get(task.priority, task.priority),
            TASK_STATUS_LABELS.get(task.status, task.status),
            task.start_date.strftime('%d/%m/%Y') if task.start_date else '',
            task.due_date.strftime('%d/%m/%Y') if task.due_date else '',
            task.created_at.strftime('%d/%m/%Y') if task.created_at else '',
        ]
        fill_color = prio_colors.get(task.priority, 'FFFFFF')
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row_n, col, val)
            c.font      = Font(name='Calibri', color='1D2A44')
            c.fill      = PatternFill('solid', fgColor=fill_color)
            c.border    = border
            c.alignment = Alignment(vertical='center', indent=1)
        ws.row_dimensions[row_n].height = 17

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── API webhook per task ID (Power Automate — più semplice del token) ──────────
@tasks_bp.post('/api/id/<int:task_id>/update')
def api_update_by_id(task_id: int):
    """Endpoint via ID numerico per Power Automate."""
    api_key = current_app.config.get('TASKS_API_KEY', '')
    if api_key and request.headers.get('X-API-Key') != api_key:
        return jsonify({'error': 'unauthorized'}), 401

    task = Task.query.get(task_id)
    if not task:
        return jsonify({'error': 'task not found'}), 404

    body = request.get_json(silent=True) or {}
    if not body and request.form:
        body = request.form.to_dict(flat=True)
    keyword = _safe_text(body.get('keyword') or body.get('status')).upper()
    note = _safe_text(body.get('note'))
    replied_by = _safe_text(body.get('replied_by'))
    email = _safe_text(body.get('email') or replied_by).lower()
    response_attachments = _parse_attachment_payload_from_json(body)

    keyword_map = {
        'PRESO IN CARICO': 'in_corso', 'PRESOINCARICO': 'in_corso',
        'IN CORSO': 'in_corso',        'INCORSO':       'in_corso',
        'BLOCCATO': 'in_attesa',       'IN ATTESA':     'in_attesa',
        'COMPLETATO': 'completato',    'ANNULLATO':     'annullato',
    }
    new_status = keyword_map.get(keyword, '')
    if not new_status and keyword.lower() in TASK_STATUS_ORDER:
        new_status = keyword.lower()

    if not new_status:
        return jsonify({
            'error': f'Parola chiave "{keyword}" non riconosciuta.',
            'valori_accettati': list(keyword_map.keys()) + TASK_STATUS_ORDER,
        }), 400

    result = _upsert_recipient_response(task, email or 'sconosciuto', new_status, note, replied_by)
    response_event = result.get('response_event')
    if response_attachments:
        _add_response_attachments(task, response_event, response_attachments,
                                  actor_name=replied_by or email, actor_email=email)
    db.session.commit()

    event_note = note
    attachment_lines = _build_attachment_lines_for_note(response_attachments)
    if attachment_lines:
        event_note = (event_note + '\n' if event_note else '') + '\n'.join(attachment_lines)

    try:
        send_task_notification(
            to_email=current_app.config['ADMIN_EMAIL'],
            subject=f'↩ [TASK-{task.id}] risposta da {replied_by or email}',
            tasks=[task],
            template_type='external_reply',
            extra={'replied_by': replied_by or email, 'note': event_note,
                   'new_status': new_status, 'new_task_status': result['new_task_status'],
                   'attachments': response_attachments},
        )
    except Exception:
        pass

    _notify_task_creator_update(task, replied_by or email, email, new_status, event_note, response_attachments)
    _safe_telegram_task_update(
        task,
        action='Aggiornamento Power App / keyword',
        actor_name=replied_by or email,
        actor_email=email,
        new_status=result.get('new_task_status', task.status),
        note=event_note,
        source='powerapp',
        extra_lines=[f'Stato destinatario: {new_status}', f'Avanzamento: {result.get("progress_percent")}% ' if result.get('progress_percent') is not None else '', *_build_attachment_lines_for_note(response_attachments)],
    )

    return jsonify({
        'ok': True,
        'task_id': task.id,
        'recipient_status': new_status,
        'task_status': task.status,
        'aggregate_changed': result['aggregate_changed'],
        'progress_percent': result['progress_percent'],
        'label': TASK_STATUS_LABELS.get(new_status, new_status),
    })


# ── API per Power App — lista task per email utente ───────────────────────────
@tasks_bp.get('/api/tasks-for-user')
def api_tasks_for_user():
    """
    Restituisce i task aperti associati all'email dell'utente.
    Usa fuzzy matching per gestire mismatch tra email Microsoft e Gmail.
    Restituisce anche resolved_email: l'email canonica trovata nelle categorie.
    """
    api_key = current_app.config.get('TASKS_API_KEY', '')
    if api_key and request.headers.get('X-API-Key') != api_key:
        return jsonify({'error': 'unauthorized'}), 401

    email = (request.args.get('email') or '').strip().lower()
    if not email or '@' not in email:
        return jsonify({'error': 'parametro email mancante o non valido'}), 400

    all_cats = TaskCategory.query.filter_by(notify_on_create=True).all()

    # Raccoglie tutte le email dei destinatari in tutte le categorie
    all_recipients = list({r.strip().lower() for c in all_cats for r in c.get_recipients()})

    # Risolve l'email canonica (gestisce mismatch Microsoft/Gmail)
    resolved = _resolve_email_in_pool(email, all_recipients)
    if resolved != email:
        current_app.logger.info('tasks-for-user: %s → %s (fuzzy match)', email, resolved)

    # Filtra le categorie che contengono la resolved email
    cat_ids = [c.id for c in all_cats if resolved in [r.strip().lower() for r in c.get_recipients()]]

    if not cat_ids:
        return jsonify({'tasks': [], 'email': email, 'resolved_email': resolved, 'count': 0})

    tasks = Task.query.filter(
        Task.category_id.in_(cat_ids),
        Task.status.notin_(['completato', 'annullato']),
    ).order_by(
        db.case({'critica': 0, 'alta': 1, 'media': 2, 'bassa': 3}, value=Task.priority),
        Task.due_date.asc().nullslast(),
    ).all()

    return jsonify({
        'tasks':          [_task_to_dict(t) for t in tasks],
        'email':          email,
        'resolved_email': resolved,
        'count':          len(tasks),
    })


@tasks_bp.get('/api/task/<int:task_id>')
def api_task_by_id(task_id: int):
    """Restituisce un singolo task per ID."""
    api_key = current_app.config.get('TASKS_API_KEY', '')
    if api_key and request.headers.get('X-API-Key') != api_key:
        return jsonify({'error': 'unauthorized'}), 401

    task = Task.query.get(task_id)
    if not task:
        return jsonify({'error': f'Task #{task_id} non trovato.'}), 404

    return jsonify({'task': _task_to_dict(task), 'found': True})


@tasks_bp.get('/api/task/by-token/<token>')
def api_task_by_token(token: str):
    api_key = current_app.config.get('TASKS_API_KEY', '')
    if api_key and request.headers.get('X-API-Key') != api_key:
        return jsonify({'error': 'unauthorized'}), 401

    task = Task.query.filter_by(external_token=token).first()
    if not task:
        return jsonify({'error': 'task not found'}), 404
    return jsonify({'task': _task_to_dict(task), 'found': True})


@tasks_bp.get('/api/task/<int:task_id>/responses')
def api_task_responses(task_id: int):
    """
    Endpoint diagnostico: mostra risposte ricevute + testa la risoluzione email.
    Uso: GET /tasks/api/task/42/responses?email=test@test.com
    """
    api_key = current_app.config.get('TASKS_API_KEY', '')
    if api_key and request.headers.get('X-API-Key') != api_key:
        return jsonify({'error': 'unauthorized'}), 401

    task = Task.query.get(task_id)
    if not task:
        return jsonify({'error': 'task not found'}), 404

    recipients = task._recipients()
    from ..models import TaskRecipientResponse
    responses  = TaskRecipientResponse.query.filter_by(task_id=task_id).all()

    # Se passata un'email di test, mostra come verrebbe risolta
    test_email    = request.args.get('email', '')
    test_resolved = None
    if test_email:
        test_resolved = _resolve_recipient_email(task, test_email)

    return jsonify({
        'task_id':           task_id,
        'task_title':        task.title,
        'task_status':       task.status,
        'progress_percent':  task.progress_percent,
        'category_recipients': recipients,
        'responses_received': [
            {
                'email':      r.recipient_email,
                'status':     r.status,
                'replied_by': r.replied_by,
                'note':       r.note,
                'updated_at': r.updated_at.isoformat() if r.updated_at else None,
                'matches_recipient': (r.recipient_email or '').strip().lower() in [x.strip().lower() for x in recipients],
            }
            for r in responses
        ],
        'missing_responses': [
            r for r in recipients
            if r.lower() not in {(resp.recipient_email or '').strip().lower() for resp in responses}
        ],
        'test_email_resolution': {
            'incoming': test_email,
            'resolved': test_resolved,
            'would_match': test_resolved in [r.strip().lower() for r in recipients],
        } if test_email else None,
    })


# ── API per Power App — aggiorna task (risposta del destinatario) ─────────────
@tasks_bp.post('/api/task-update')
def api_task_update_powerapp():
    """Endpoint principale Power App/Power Automate."""
    api_key = current_app.config.get('TASKS_API_KEY', '')
    if api_key and request.headers.get('X-API-Key') != api_key:
        return jsonify({'ok': False, 'message': 'Non autorizzato.'}), 401

    body = request.get_json(silent=True) or {}
    if not body and request.form:
        body = request.form.to_dict(flat=True)
    task_id = _normalize_task_id(body.get('task_id'))
    task_token = _safe_text(body.get('task_token') or body.get('token') or body.get('external_token'))
    new_status = _safe_text(body.get('status')).lower()
    note = _safe_text(body.get('note'))
    replied_by = _safe_text(body.get('replied_by'))
    email = _safe_text(body.get('email') or replied_by).lower()
    response_attachments = _parse_attachment_payload_from_json(body)

    task = None
    if task_id is not None:
        task = Task.query.get(task_id)
    if not task and task_token:
        task = Task.query.filter_by(external_token=task_token).first()
    if not task:
        return jsonify({'ok': False, 'message': 'Task non trovato. Aggiorna la pagina e riprova.'}), 404

    if new_status not in TASK_STATUS_ORDER:
        return jsonify({
            'ok': False,
            'message': f'Stato "{new_status}" non valido.',
            'valori_accettati': TASK_STATUS_ORDER,
        }), 400

    result = _upsert_recipient_response(task, email or replied_by or 'sconosciuto',
                                        new_status, note, replied_by)
    response_event = result.get('response_event')
    if response_attachments:
        _add_response_attachments(task, response_event, response_attachments,
                                  actor_name=replied_by or email, actor_email=email)
    db.session.commit()

    event_note = note
    attachment_lines = _build_attachment_lines_for_note(response_attachments)
    if attachment_lines:
        event_note = (event_note + '\n' if event_note else '') + '\n'.join(attachment_lines)

    try:
        send_task_notification(
            to_email=current_app.config['ADMIN_EMAIL'],
            subject=f'↩ [TASK-{task.id}] risposta da {replied_by or email}',
            tasks=[task],
            template_type='external_reply',
            extra={'replied_by': replied_by or email, 'note': event_note,
                   'new_status': new_status,
                   'aggregate_changed': result['aggregate_changed'],
                   'new_task_status': result['new_task_status'],
                   'attachments': response_attachments},
        )
    except Exception:
        pass

    _notify_task_creator_update(task, replied_by or email, email, new_status, event_note, response_attachments)
    _safe_telegram_task_update(
        task,
        action='Aggiornamento Power App',
        actor_name=replied_by or email,
        actor_email=email,
        new_status=result.get('new_task_status', task.status),
        note=event_note,
        source='powerapp',
        extra_lines=[f'Stato destinatario: {new_status}', f'Avanzamento: {result.get("progress_percent")}% ' if result.get('progress_percent') is not None else '', *_build_attachment_lines_for_note(response_attachments)],
    )

    status_label = TASK_STATUS_LABELS.get(new_status, new_status)
    return jsonify({
        'ok': True,
        'task_id': task.id,
        'task_title': task.title,
        'recipient_status': new_status,
        'status_label': status_label,
        'task_status': task.status,
        'task_status_label': TASK_STATUS_LABELS.get(task.status, task.status),
        'aggregate_changed': result['aggregate_changed'],
        'progress_percent': result['progress_percent'],
        'attachments': [_serialize_attachment(att) for att in task.attachments.filter_by(attachment_scope='response').all()],
        'message': f'Risposta registrata: {status_label}. ' + (
            f'Task avanzato a "{TASK_STATUS_LABELS.get(task.status, task.status)}".'
            if result['aggregate_changed'] else
            f'Task ancora in "{TASK_STATUS_LABELS.get(task.status, task.status)}" ' +
            (f'({result["progress_percent"]}% avanzati)' if result['progress_percent'] else '')
        ),
    })
