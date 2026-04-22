from __future__ import annotations
from copy import copy
from pathlib import Path
import shutil

from openpyxl import load_workbook
from .data_models import WeeklyProgramData


class WeeklyProgramGenerationError(RuntimeError):
    pass


class WeeklyProgramTemplateService:
    DATE_HEADER_CELLS = ['B5', 'C5', 'D5', 'E5', 'F5']
    LOCATION_CELLS    = ['B6', 'C6', 'D6', 'E6', 'F6']
    BLACK = 'FF000000'

    def __init__(self, template_path: str | Path):
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f'Template non trovato: {self.template_path}')

    def generate(self, data: WeeklyProgramData, target_dir: str | Path) -> dict[str, str]:
        target_dir = Path(target_dir)
        target_dir.mkdir(parents=True, exist_ok=True)
        output_xlsx = target_dir / f'{data.safe_filename_base}.xlsx'
        shutil.copy2(self.template_path, output_xlsx)

        try:
            wb = load_workbook(output_xlsx)
            ws = wb.active

            if data.start_date and data.end_date:
                ws['A1'] = (f'DAL  {data.start_date.strftime("%d/%m/%Y")}'
                            f'   al {data.end_date.strftime("%d/%m/%Y")}')
            elif data.start_date:
                ws['A1'] = f'DAL  {data.start_date.strftime("%d/%m/%Y")}'

            ws['A6'] = data.full_name
            self._black(ws['A1'])
            self._black(ws['A6'])

            for cell_ref, dt in zip(self.DATE_HEADER_CELLS, data.effective_dates):
                ws[cell_ref] = dt
                ws[cell_ref].number_format = 'DD/MM/YYYY'
                self._black(ws[cell_ref])

            for cell_ref, loc in zip(self.LOCATION_CELLS, data.day_locations):
                ws[cell_ref] = loc
                self._black(ws[cell_ref])

            wb.save(output_xlsx)
        except Exception as exc:
            raise WeeklyProgramGenerationError(str(exc)) from exc

        return {'xlsx': str(output_xlsx)}

    def _black(self, cell) -> None:
        f = copy(cell.font)
        f.color = self.BLACK
        cell.font = f
